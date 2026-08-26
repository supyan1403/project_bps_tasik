import os
import re
import csv
import io
import json
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")))
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import models
from database import get_db
from pipeline import parse_indonesian_number
from routers.tables import (
    get_table_headers, 
    clean_bilingual_header, 
    get_safe_windows_path, 
    normalize_record_first_col
)

router = APIRouter(prefix="/api", tags=["Time Series Analysis"])

def extract_timeseries_year(year_val, default_year=None):
    if year_val is None or year_val == '':
        return default_year
    s = str(year_val).strip()
    m = re.search(r'\b(19\d{2}|20\d{2})\b', s)
    if m:
        return int(m.group(1))
    return default_year

def _clean_header_for_master(header: str) -> str:
    if not header:
        return ""
    h = clean_bilingual_header(header)
    h = re.sub(r'\.\d+$', '', h)
    return h.strip()

def get_column_years_from_db(db: Session, table, doc_year: int) -> dict:
    headers = get_table_headers(db, table)
    if not headers:
        return {}
    years = table.years or []
    col_years = {}

    first_lower = headers[0].lower() if headers else ""
    is_vertical = "tahun" in first_lower or "year" in first_lower

    if is_vertical:
        rows = db.query(models.TableRow).filter(models.TableRow.table_id == table.id).all()
        years_in_col0 = set()
        for r in rows:
            val = str(r.data.get(headers[0], "")).strip()
            y_extracted = extract_timeseries_year(val)
            if y_extracted:
                years_in_col0.add(y_extracted)
        for idx, h in enumerate(headers):
            if idx == 0:
                continue
            if not any(char.isalpha() for char in h):
                continue
            col_years.setdefault(_clean_header_for_master(h), set()).update(years_in_col0)
        col_years = {k: sorted(list(v)) for k, v in col_years.items()}
    else:
        for idx, h in enumerate(headers):
            if idx == 0:
                continue
            if not any(char.isalpha() for char in h):
                continue
            yr_val = years[idx] if idx < len(years) else ""
            col_year = extract_timeseries_year(yr_val, default_year=doc_year - 1)
            col_years.setdefault(_clean_header_for_master(h), set()).add(col_year)
        col_years = {k: sorted(list(v)) for k, v in col_years.items()}

    return col_years

def get_column_years_with_position(db: Session, table, doc_year: int) -> dict:
    headers = get_table_headers(db, table)
    if not headers:
        return {}
    years = table.years or []
    result = {}
    first_lower = headers[0].lower() if headers else ""
    is_vertical = "tahun" in first_lower or "year" in first_lower

    if is_vertical:
        rows = db.query(models.TableRow).filter(models.TableRow.table_id == table.id).all()
        years_in_col0 = set()
        for r in rows:
            val = str(r.data.get(headers[0], "")).strip()
            y_extracted = extract_timeseries_year(val)
            if y_extracted:
                years_in_col0.add(y_extracted)
        for idx, h in enumerate(headers):
            if idx == 0:
                continue
            if not any(char.isalpha() for char in h):
                continue
            name = _clean_header_for_master(h)
            result.setdefault(name, {"years": set(), "pos": idx})
            result[name]["years"].update(years_in_col0)
    else:
        for idx, h in enumerate(headers):
            if idx == 0:
                continue
            if not any(char.isalpha() for char in h):
                continue
            name = _clean_header_for_master(h)
            yr_val = years[idx] if idx < len(years) else ""
            col_year = extract_timeseries_year(yr_val, default_year=doc_year - 1)
            result.setdefault(name, {"years": set(), "pos": idx})
            result[name]["years"].add(col_year)

    return result

def get_clean_chapter_name(level1: str) -> str:
    CHAPTER_NAMES = {
        "1": "Geografi dan Iklim",
        "2": "Pemerintahan",
        "3": "Penduduk dan Ketenagakerjaan",
        "4": "Sosial dan Kesejahteraan Rakyat",
        "5": "Pertanian, Kehutanan, dan Perikanan",
        "6": "Industri, Pertambangan, Energi, dan Air",
        "7": "Pariwisata",
        "8": "Transportasi dan Komunikasi",
        "9": "Koperasi dan Usaha Mikro Kecil Menengah (UMKM)",
        "10": "Pengeluaran dan Konsumsi Penduduk",
        "11": "Perdagangan",
        "12": "Pendapatan Regional",
        "13": "Perbandingan Regional / Antar Wilayah"
    }
    return CHAPTER_NAMES.get(str(level1).strip(), f"Bab {level1}")

def get_clean_table_name(table_name: str) -> str:
    if not table_name:
        return ""
    name = re.sub(r'^(?:Tabel[\s_]*\d+(?:\.\d+)*\s*(?:-\s*|:\s*|)\s*)', '', table_name, flags=re.IGNORECASE)
    name = re.sub(r'\.csv$', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s*\((?:Hal|Halaman|hlm)[\s\d,\-–—\.\?]+\)', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s*\(\s*\d+[\s,\d\-–—\.]*\)\s*$', '', name)
    year_token = r'(?:19|20)\d{2}[*\d]?'
    year_conn = r'(?:\s*(?:[-–—/]|dan|and|sd|s/d|to|,)\s*' + year_token + r')*'
    name = re.sub(r'[,.\s]+(?:(?:pada|di)\s+)?(?:tahun|years?)\s*(?:' + year_token + year_conn + r'.*)?$', '', name, flags=re.IGNORECASE)
    name = re.sub(r'[,.\s]+' + year_token + year_conn + r'.*$', '', name, flags=re.IGNORECASE)
    name = re.sub(r'[,.\s]+(?:(?:pada|di)\s+)?(?:tahun|years?)\s*$', '', name, flags=re.IGNORECASE)
    return re.sub(r'[,.\-\s–—]+$', '', name).strip()

_KECAMATAN_TASIK = [
    "cipatujah", "karangnunggal", "cikalong", "pancatengah", "cikatomas",
    "cibalong", "parungponteng", "bantarkalong", "bojongasih", "culamega",
    "bojonggambir", "sodonghilir", "taraju", "salawu", "puspahiang",
    "tanjungjaya", "sukaraja", "salopa", "jatiwaras", "cineam",
    "karangjaya", "manonjaya", "gunungtanjung", "singaparna", "mangunreja",
    "sukarame", "cigalontang", "leuwisari", "padakembang", "sariwangi",
    "sukaratu", "cisayong", "sukahening", "rajapolah", "jamanis",
    "ciawi", "kadipaten", "pagerageung", "sukaresik", "kabupaten tasikmalaya"
]
_KECAMATAN_SET = set(_KECAMATAN_TASIK)

def normalize_entity_name(raw: str) -> str:
    if not raw:
        return raw
    s = raw.strip()
    s = re.sub(r'^\d+[\s\.]+', '', s)
    s = re.sub(r'\s*\([^)]*\)', '', s)
    s = s.strip()
    s_lower = s.lower()
    for kec in _KECAMATAN_TASIK:
        if s_lower == kec:
            return kec.title()
    return s.title() if s.islower() else s

def normalize_entity_key(raw_key: str) -> str:
    if not raw_key:
        return "Rincian"
    k = raw_key.lower().strip()
    if any(w in k for w in ["kecamatan", "distrik"]):
        return "Kecamatan"
    if any(w in k for w in ["desa", "kelurahan"]):
        return "Desa/Kelurahan"
    if any(w in k for w in ["kabupaten", "kota"]):
        return "Kabupaten/Kota"
    if "provinsi" in k:
        return "Provinsi"
    if any(w in k for w in ["tahun", "year"]):
        return "Tahun"
    if any(w in k for w in ["bulan", "month"]):
        return "Bulan"
    return "Rincian"

def _extract_year_range_from_table_name(table_name: str):
    if not table_name:
        return None
    m = re.search(r'\b(19\d{2}|20\d{2})\s*[\-–—~s\.\/]+s*d?\.?\s*(19\d{2}|20\d{2})\b', table_name)
    if m:
        y1, y2 = int(m.group(1)), int(m.group(2))
        return (min(y1, y2), max(y1, y2))
    return None

def _get_unit_multiplier(unit: str = "", table_name: str = "", header: str = "") -> float:
    combined = f"{unit} {table_name} {header}".lower()
    if any(k in combined for k in ["per kapita", "/kapita", "per orang", "per bulan", "persen", "%", "rasio", "indeks", "kepadatan", "laju", "umur harapan"]):
        return 1.0
    if re.search(r'\b(miliar|milyar)\b', combined):
        return 1000000000.0
    elif re.search(r'\bjuta\b', combined):
        return 1000000.0
    elif re.search(r'\b(ribu|thousands)\b', combined):
        return 1000.0
    return 1.0

def _format_scaled_indo_number(val_float) -> str:
    if val_float is None:
        return ""
    if val_float == int(val_float):
        n = int(val_float)
        return '{:,}'.format(n).replace(',', '.')
    else:
        fixed = f"{val_float:.2f}"
        int_part, dec_part = fixed.split('.')
        int_str = '{:,}'.format(int(int_part)).replace(',', '.')
        dec_clean = dec_part.rstrip('0')
        return f"{int_str},{dec_clean}" if dec_clean else int_str

def _normalize_base_unit(unit_str: str) -> str:
    if not unit_str:
        return unit_str
    u = unit_str.strip()
    u_lower = u.lower()
    if u_lower in ['ribu', 'ribu jiwa', '(ribu)', '(ribu jiwa)']:
        return 'jiwa'
    if u_lower in ['ribu orang', '(ribu orang)']:
        return 'orang'
    if u_lower in ['ribu rupiah', 'juta rupiah', 'miliar rupiah']:
        return 'rupiah'
    if u_lower in ['ribu ton']:
        return 'ton'
    if u_lower in ['ribu ha']:
        return 'ha'
    if u_lower.startswith('ribu '):
        return u[5:].strip()
    return u

def _normalize_indo_number(val: str, unit: str = "", table_name: str = "", header: str = "") -> str:
    if not val:
        return val
    s = str(val).strip()
    if not s or s in ['-', '...', '']:
        return s
    
    num = parse_indonesian_number(s)
    if num is None:
        return s
    
    mult = _get_unit_multiplier(unit, table_name, header)
    if mult == 1000.0:
        if num >= 60000.0:
            mult = 1.0
        elif " " in s and len(s.replace(" ", "")) >= 6:
            mult = 1.0
        elif re.match(r'^[1-9]\d{0,2}\.\d{3}$', s):
            mult = 1.0
        elif s.count('.') > 1:
            parts = s.split('.')
            if len(parts[-1]) == 3:
                mult = 1.0
    elif mult == 1000000.0 and num >= 1000000.0:
        mult = 1.0
        
    scaled = num * mult
    return _format_scaled_indo_number(scaled)

def _extract_unit_from_table_name(table_name: str) -> str:
    if not table_name:
        return ""
    matches = re.findall(r'\(([^)]+)\)', table_name)
    _NON_UNIT_ABBREVS = {'apk', 'apm', 'asn', 'iumk', 'tpak', 'tpt', 'ra', 'mi', 'ma', 'mts', 'sd', 'smp', 'sma', 'smk', 'tk', 'persero'}
    for m in matches:
        m_stripped = m.strip()
        m_lower = m_stripped.lower()
        if m_lower.startswith('hal') or re.match(r'^\d{4}', m_lower) or m_lower in _NON_UNIT_ABBREVS or len(m_stripped) > 30:
            continue
        return m_stripped
    return ""

def _infer_unit_from_indicator(indicator_name: str) -> str:
    if not indicator_name:
        return ""
    name_lower = indicator_name.lower().strip()
    DIMENSION_PATTERNS = [
        r'^(kecamatan|kabupaten|kota|provinsi|negara|ibukota|ibukota\s+kecamatan)$',
        r'^(jenis\s+tanaman|jenis\s+ikan|jenis\s+industri|jenis\s+koperasi|jenis\s+pengeluaran|jenis\s+pendapatan|jenis\s+permukaan|jenis\s+penangkapan|jenis\s+pengelolaan|jenis\s+sarana|jenis\s+pelanggan)$',
        r'^(lapangan\s+usaha|subsektor|sektor)$',
        r'^(kelompok\s+umur|kelompok\s+komoditas|golongan|golongan\s+tarip|golongan\s+tarif|pangkat|jenjang|tingkat\s+pendidikan|jabatan|kegiatan\s+utama|status\s+pekerjaan)$',
        r'^(partai\s+politik|partai)$',
        r'^(uraian|rincian|kategori|keterangan|kondisi\s+jalan|objek\s+wisata)$',
        r'^(triw|bulan|tahun)$',
        r'^(menurut\s+kabupaten|menurut\s+kecamatan)$',
        r'^(tingkat\s+kewenangan)$'
    ]
    for pat in DIMENSION_PATTERNS:
        if re.search(pat, name_lower):
            return ""
    _UNIT_KEYWORD_MAP = [
        (r'kepadatan\s+penduduk', 'jiwa/km²'),
        (r'rasio\s+jenis\s+kelamin', 'rasio'),
        (r'indeks\s+pembangunan\s+manusia', 'indeks'),
        (r'angka\s+partisipasi', '%'),
        (r'tingkat\s+pengangguran', '%'),
        (r'laju\s+pertumbuhan', '%'),
        (r'persentase', '%'),
        (r'populasi\s+(ternak|unggas)', 'ekor'),
        (r'produksi\s+(daging|susu|telur|ikan|budidaya|perikanan|padi|kedelai|jagung|ubi|tebu|tembakau)', 'ton'),
        (r'produksi\s+(mangga|durian|jeruk|pisang|pepaya|salak|manggis|alpukat|bawang|cabai|kentang|kubis|tomat|sayuran|buah)', 'kuintal'),
        (r'luas\s+(panen|areal|lahan|sawah|hutan|tanaman)', 'ha'),
        (r'luas\s+(wilayah|daerah)', 'km²'),
        (r'jumlah\s+(guru|murid|dosen|siswa|dokter|perawat|bidan|tenaga\s+kesehatan|anggota|karyawan)', 'orang'),
        (r'jumlah\s+penduduk', 'jiwa'),
        (r'jumlah\s+(puskesmas|posyandu|masjid|mushola|gereja|pura|vihara|hotel|koperasi|pasar|industri|kantor|sekolah)', 'unit'),
        (r'pengeluaran\s+per\s+kapita', 'rupiah'),
        (r'upah\s+minimum', 'rupiah'),
        (r'produk\s+domestik\s+regional\s+bruto', 'juta rupiah'),
        (r'panjang\s+(jalan|saluran|irigasi)', 'km'),
        (r'curah\s+hujan', 'mm'),
        (r'suhu\s+rata-rata', '°C'),
    ]
    for pattern, unit in _UNIT_KEYWORD_MAP:
        if re.search(pattern, name_lower):
            return unit
    return ""

def _infer_unit_from_headers_and_table(headers: list, table_name: str) -> str:
    unit = _extract_unit_from_table_name(table_name)
    if unit:
        return _normalize_base_unit(unit)
    for h in headers:
        u = _infer_unit_from_indicator(h)
        if u:
            return _normalize_base_unit(u)
    return ""

def _build_vk_units(headers: list, table_name: str) -> dict:
    table_unit = _extract_unit_from_table_name(table_name)
    result = {}
    for h in headers:
        u = _infer_unit_from_indicator(h) or table_unit or ""
        result[h] = _normalize_base_unit(u)
    return result

def _classify_entity_type(entity_name: str) -> str:
    if not entity_name:
        return "Lainnya"
    name_lower = entity_name.lower().strip()
    if name_lower.startswith("kab.") or name_lower.startswith("kabupaten") or name_lower.startswith("kota"):
        return "Kabupaten/Kota"
    if "provinsi" in name_lower or name_lower in ["jawa barat", "jawa tengah", "jawa timur", "banten", "dki jakarta"]:
        return "Provinsi"
    if name_lower in _KECAMATAN_SET:
        return "Kecamatan"
    if name_lower in ["indonesia", "asia", "eropa", "amerika", "dunia", "world"]:
        return "Nasional/Internasional"
    if name_lower in ["jumlah", "total", "subtotal", "grand total", "keseluruhan", "seluruh"]:
        return "Total"
    return "Lainnya"

def check_cell_format_anomaly(raw_val: str, prev_raw_val: str = None) -> Optional[str]:
    s = str(raw_val).strip()
    if not s or s in ["-", "...", "–", "—", ""]:
        return None
    if "?" in s or ".." in s or ",," in s:
        return f"Karakter/simbol rusak pada angka ({s})"
    if re.search(r'^\d{1,3}\s+\d{3}', s):
        return f"Spasi pemisah ribuan janggal ({s})"
    
    p = str(prev_raw_val).strip() if prev_raw_val else ""
    p_valid = bool(p) and p not in ["-", "...", "–", "—", ""]
    _C3 = r'^[1-9]\d{0,2},\d{3}$'
    _P3 = r'^\d{1,3}\.\d{3}$'
    _PD = r'^\d+\.\d{1,2}$'
    _CD = r'^\d+,\d{1,2}$'
    
    if p_valid:
        if re.match(_P3, p) and re.match(_C3, s):
            return f"Inkonsistensi format: tahun sebelumnya titik ribuan ({p}), tahun ini koma ({s})"
        if re.match(_C3, p) and re.match(_P3, s):
            return f"Inkonsistensi format: tahun sebelumnya koma ({p}), tahun ini titik ribuan ({s})"
        if re.match(_CD, p) and re.match(_PD, s):
            return f"Inkonsistensi format: tahun sebelumnya koma desimal ({p}), tahun ini titik ({s})"
        if re.match(_PD, p) and re.match(_CD, s):
            return f"Inkonsistensi format: tahun sebelumnya titik desimal ({p}), tahun ini koma ({s})"
        if (re.match(_PD, p) and re.match(_PD, s)) or (re.match(_CD, p) and re.match(_CD, s)) or (re.match(_P3, p) and re.match(_P3, s)) or (re.match(_C3, p) and re.match(_C3, s)):
            return None
    if re.match(_PD, s):
        return f"Format salah: menggunakan titik desimal ({s}) alih-alih koma desimal"
    return None

def detect_timeseries_anomalies(tables_data: list) -> list:
    if not tables_data:
        return []
    entity_series = {}
    for t in tables_data:
        yr = t.get("year")
        if not yr or not isinstance(yr, int):
            continue
        for row in t.get("data", []):
            ent = row.get("entitas", "").strip()
            if not ent:
                continue
            if ent not in entity_series:
                entity_series[ent] = {}
            for vk, val_str in row.get("nilai", {}).items():
                if vk not in entity_series[ent]:
                    entity_series[ent][vk] = {}
                entity_series[ent][vk][yr] = str(val_str).strip()

    anomalies = []
    for ent, vk_dict in entity_series.items():
        for vk, yr_dict in vk_dict.items():
            sorted_years = sorted(yr_dict.keys())
            if len(sorted_years) < 2:
                continue
            for i in range(1, len(sorted_years)):
                prev_yr = sorted_years[i - 1]
                curr_yr = sorted_years[i]
                if curr_yr - prev_yr > 3:
                    continue
                prev_raw = yr_dict[prev_yr]
                curr_raw = yr_dict[curr_yr]
                err = check_cell_format_anomaly(curr_raw, prev_raw)
                if err:
                    anomalies.append({
                        "entitas": ent,
                        "indicator": vk,
                        "year": curr_yr,
                        "prev_year": prev_yr,
                        "val": curr_raw,
                        "prev_val": prev_raw,
                        "type": "format",
                        "severity": "high",
                        "message": f"Nilai '{ent}' ({vk}): {err}."
                    })
                pnum = parse_indonesian_number(prev_raw)
                cnum = parse_indonesian_number(curr_raw)
                if pnum and cnum and pnum != 0 and cnum != 0:
                    ratio = cnum / pnum
                    if abs(ratio) > 100 or abs(ratio) < 0.01:
                        scale_desc = f"naik {ratio:.0f}x" if ratio > 1 else f"turun ke 1/{1/ratio:.0f}-nya"
                        anomalies.append({
                            "entitas": ent,
                            "indicator": vk,
                            "year": curr_yr,
                            "prev_year": prev_yr,
                            "val": curr_raw,
                            "prev_val": prev_raw,
                            "type": "scale",
                            "severity": "medium",
                            "message": f"Nilai '{ent}' ({vk}): {scale_desc} drastis antara {prev_yr} ({prev_raw}) dan {curr_yr} ({curr_raw})."
                        })
    return anomalies

def _dedup_timeseries_results(results):
    raw_groups = {}
    for r in results:
        key = (r["year"], r.get("entity_key", "").strip().lower())
        raw_groups.setdefault(key, []).append(r)

    groups = {}
    for key, group in raw_groups.items():
        subgroups = []
        for r in group:
            rset = set(r.get("headers", []))
            merged = False
            for i, (g_headers, gset, members) in enumerate(subgroups):
                if rset == gset or rset.issubset(gset) or gset.issubset(rset):
                    if len(rset) > len(gset):
                        subgroups[i] = (r["headers"], rset, members + [r])
                    else:
                        subgroups[i] = (g_headers, gset, members + [r])
                    merged = True
                    break
            if not merged:
                subgroups.append((r["headers"], rset, [r]))
        groups[key] = subgroups

    deduped = []
    for (year, entity_key), subgroups in groups.items():
        for headers_tuple, hset, group in subgroups:
            group.sort(key=lambda r: r.get("doc_year", 0), reverse=True)
            merged_data = {}
            merged_sources = {}
            best_table_info = group[0]

            for r in group:
                for k, sinfo in r.get("sources", {}).items():
                    if k not in merged_sources:
                        merged_sources[k] = sinfo
                for row in r["data"]:
                    ent = row["entitas"]
                    if ent not in merged_data:
                        merged_data[ent] = {**row, "nilai": dict(row.get("nilai", {})), "sumber": dict(row.get("sumber", {}))}
                    else:
                        for k, v in row.get("nilai", {}).items():
                            cur = merged_data[ent]["nilai"].get(k)
                            if v not in (None, "") and cur in (None, ""):
                                merged_data[ent]["nilai"][k] = v
                                if "sumber" in row and k in row["sumber"]:
                                    merged_data[ent].setdefault("sumber", {})[k] = row["sumber"][k]

            if merged_data:
                data_rows = list(merged_data.values())
                data_rows.sort(key=lambda r: (r["entitas"] == "Kabupaten Tasikmalaya", r["entitas"]))
                deduped.append({
                    "table_id": best_table_info["table_id"],
                    "table_name": best_table_info["table_name"],
                    "year": year,
                    "doc_year": best_table_info.get("doc_year", 0),
                    "doc_filename": best_table_info.get("doc_filename", ""),
                    "entity_key": best_table_info["entity_key"],
                    "headers": list(headers_tuple),
                    "unit": best_table_info.get("unit", ""),
                    "vk_units": best_table_info.get("vk_units", {}),
                    "sources": merged_sources,
                    "data": data_rows
                })
    return deduped

class TimeSeriesExportRequest(BaseModel):
    years: List[int]
    valueKeys: List[str]
    vkUnits: Optional[Dict[str, str]] = {}
    entities: Optional[List[str]] = []
    entityMap: Dict[str, Any]

@router.get("/timeseries/catalog")
def timeseries_catalog(db: Session = Depends(get_db)):
    rows = db.query(models.ExtractedTable.id, models.ExtractedTable.table_name, models.Document.year).join(
        models.Document, models.ExtractedTable.document_id == models.Document.id
    ).all()
    seen = set()
    items = []
    for row in rows:
        key = (row.table_name.strip() if row.table_name else '', row.year)
        if key in seen:
            continue
        seen.add(key)
        name = key[0]
        m = re.match(r'(?:tabel\s*)?(\d+(?:\.\d+)*)', name, re.I)
        parts = m.group(1).split('.') if m else []
        level1 = parts[0] if len(parts) >= 1 else ''
        level2 = parts[0] + '.' + parts[1] if len(parts) >= 2 else ''
        table_prefix = f"Tabel {m.group(1)}" if m else ""
        items.append({
            "table_id": row.id,
            "table_name": name,
            "clean_table_name": get_clean_table_name(name),
            "table_prefix": table_prefix,
            "level1": level1,
            "clean_chapter_name": get_clean_chapter_name(level1),
            "level2": level2,
            "year": row.year
        })
    return {"status": "success", "data": items}

@router.get("/timeseries/table-details")
def timeseries_table_details(table_prefix: str, db: Session = Depends(get_db)):
    prefix_clean = table_prefix.strip()
    tables = db.query(models.ExtractedTable).filter(models.ExtractedTable.table_name.like(f"%{prefix_clean}%")).all()
    if not tables:
        return {"status": "success", "indicators": []}
    doc_ids = [t.document_id for t in tables]
    docs = db.query(models.Document).filter(models.Document.id.in_(doc_ids)).all()
    doc_year_map = {d.id: d.year for d in docs}
    all_indicators = {}
    for table in tables:
        doc_year = doc_year_map.get(table.document_id, 2026)
        col_years = get_column_years_from_db(db, table, doc_year)
        for col_name, yrs in col_years.items():
            all_indicators.setdefault(col_name, set()).update(yrs)
    result = [{"name": col_name, "years": sorted(list(yrs))} for col_name, yrs in all_indicators.items()]
    return {"status": "success", "indicators": result}

@router.get("/timeseries/indicator-years")
def timeseries_indicator_years(db: Session = Depends(get_db)):
    tables = db.query(models.ExtractedTable).all()
    if not tables:
        return {"status": "success", "indicators": []}
    doc_ids = [t.document_id for t in tables]
    docs = db.query(models.Document).filter(models.Document.id.in_(doc_ids)).all()
    doc_year_map = {d.id: d.year for d in docs}

    def _extract_bab(tname):
        if not tname: return None
        m = re.search(r'Tabel[\s_]*(\d+)', tname, re.I) or re.search(r'(\d+)', tname)
        return int(m.group(1)) if m else None

    tables_ordered = sorted(tables, key=lambda t: (_extract_bab(t.table_name) or 9999, doc_year_map.get(t.document_id, 9999), t.id))
    all_indicators = {}
    for table in tables_ordered:
        doc_year = doc_year_map.get(table.document_id, 2026)
        bab = _extract_bab(table.table_name)
        col_data = get_column_years_with_position(db, table, doc_year)
        for name, info in col_data.items():
            if name not in all_indicators:
                all_indicators[name] = {"years": set(), "order": [bab if bab is not None else 9999, info["pos"]]}
            all_indicators[name]["years"].update(info["years"])

    result = []
    for name, info in all_indicators.items():
        raw_bab = info["order"][0] if info["order"][0] != 9999 else None
        bab_name = get_clean_chapter_name(str(raw_bab)) if raw_bab is not None else None
        if bab_name and bab_name == f"Bab {raw_bab}":
            bab_name = None
        result.append({
            "name": name,
            "years": sorted(list(info["years"])),
            "bab_num": raw_bab,
            "bab_name": bab_name,
            "order": info["order"]
        })
    result.sort(key=lambda x: (x["order"][0], x["order"][1], x["name"]))
    return {"status": "success", "indicators": result}

@router.get("/timeseries/data-by-indicators")
def timeseries_data_by_indicators(indicators: str = "", years: str = "", db: Session = Depends(get_db)):
    """Fetch multi-year time series table data by selected indicator names and years."""
    if not indicators or not years:
        return {"status": "success", "data": []}

    ind_list = [i.strip() for i in indicators.split(",") if i.strip()]
    ind_set = {i.lower() for i in ind_list}
    
    year_list = []
    for y in years.split(","):
        y_str = y.strip()
        if y_str.isdigit():
            year_list.append(int(y_str))
    year_set = set(year_list)
    
    if not ind_set or not year_set:
        return {"status": "success", "data": []}

    tables = db.query(models.ExtractedTable).all()
    doc_ids = [t.document_id for t in tables]
    docs = db.query(models.Document).filter(models.Document.id.in_(doc_ids)).all()
    doc_map = {d.id: d for d in docs}

    matched_results = []
    for table in tables:
        doc = doc_map.get(table.document_id)
        doc_year = doc.year if doc else 2026
        col_years = get_column_years_from_db(db, table, doc_year)

        matched_cols = [c for c in col_years if c.strip().lower() in ind_set]
        if not matched_cols:
            continue

        all_rows = db.query(models.TableRow).filter(models.TableRow.table_id == table.id).all()
        if not all_rows:
            continue

        headers = list(all_rows[0].data.keys()) if all_rows[0].data else []
        entity_key = headers[0] if headers else "Kecamatan"
        
        # Filter hanya kolom indikator yang dicentang/diminta oleh user
        value_cols = [c for c in headers if c != entity_key and any(c.strip().lower() == mc.strip().lower() for mc in matched_cols)]
        if not value_cols:
            continue

        for col in matched_cols:
            for yr in col_years[col]:
                if yr in year_set:
                    data_rows = []
                    for r in all_rows:
                        record = r.data
                        ent = str(record.get(entity_key, "")).strip()
                        if not ent or ent in ["-", "..."]:
                            continue
                        ent = normalize_entity_name(ent)
                        vals = {c: _normalize_indo_number(str(record.get(c, "")).strip(), unit="", table_name=table.table_name, header=c) for c in value_cols}
                        data_rows.append({"entitas": ent, "tipe": _classify_entity_type(ent), "nilai": vals})

                    matched_results.append({
                        "table_id": table.id,
                        "table_name": table.table_name,
                        "year": yr,
                        "doc_year": doc_year,
                        "doc_filename": doc.filename if doc else "",
                        "entity_key": entity_key,
                        "headers": value_cols,
                        "unit": _infer_unit_from_headers_and_table(value_cols, table.table_name),
                        "vk_units": _build_vk_units(value_cols, table.table_name),
                        "sources": {col: {"table_id": table.id, "table_name": table.table_name, "doc_filename": doc.filename if doc else ""}},
                        "data": data_rows
                    })

    deduped = _dedup_timeseries_results(matched_results)
    return {"status": "success", "data": deduped}

@router.get("/search/timeseries")
def search_timeseries(keyword: str = "", start_year: int = None, end_year: int = None, db: Session = Depends(get_db)):
    """Legacy keyword-based time series search wrapper."""
    if not keyword:
        return {"status": "success", "data": []}
    
    # Forward ke logic indicators
    kw_clean = keyword.strip()
    return timeseries_data_by_indicators(indicators=kw_clean, years=",".join(str(y) for y in range(start_year or 2015, (end_year or 2030) + 1)), db=db)

@router.get("/timeseries/browse-data")
def timeseries_browse_data(table_id: int, db: Session = Depends(get_db)):
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(404, "Table not found")
    doc = db.query(models.Document).filter(models.Document.id == table.document_id).first()
    year = doc.year if doc else None
    all_rows = db.query(models.TableRow).filter(models.TableRow.table_id == table_id).all()
    if not all_rows:
        return {"status": "success", "data": []}

    headers = list(all_rows[0].data.keys()) if all_rows[0].data else []
    entity_key = headers[0] if headers else "Kecamatan"
    value_cols = [c for c in headers if c != entity_key]
    data_rows = []
    for r in all_rows:
        record = r.data
        ent = str(record.get(entity_key, '')).strip()
        if not ent or ent in ['-', '...']:
            continue
        ent = normalize_entity_name(ent)
        vals = {c: _normalize_indo_number(str(record.get(c, '')).strip()) for c in value_cols}
        data_rows.append({"entitas": ent, "tipe": _classify_entity_type(ent), "nilai": vals})

    data_rows.sort(key=lambda r: (r["entitas"] == "Kabupaten Tasikmalaya", r["entitas"]))
    column_years = get_column_years_from_db(db, table, year if year else 2026)
    return {"status": "success", "data": [{
        "table_id": table.id,
        "table_name": table.table_name,
        "year": year,
        "entity_key": entity_key,
        "headers": value_cols,
        "unit": _infer_unit_from_headers_and_table(value_cols, table.table_name),
        "vk_units": _build_vk_units(value_cols, table.table_name),
        "data": data_rows,
        "column_years": column_years
    }]}

@router.get("/timeseries/table-columns")
def timeseries_table_columns(table_ids: str, db: Session = Depends(get_db)):
    ids = [int(x) for x in table_ids.split(",") if x.strip()]
    if not ids:
        return {"status": "success", "entity_key": "", "columns": []}
    tables = db.query(models.ExtractedTable).filter(models.ExtractedTable.id.in_(ids)).all()
    entity_key = ""
    all_cols = {}
    for table in tables:
        row = db.query(models.TableRow).filter(models.TableRow.table_id == table.id).first()
        if not row or not row.data:
            continue
        headers = list(row.data.keys())
        if not headers:
            continue
        if not entity_key:
            entity_key = headers[0]
        for h in headers:
            if h != entity_key:
                all_cols[h.lower().strip()] = h
    return {"status": "success", "entity_key": entity_key, "columns": sorted(all_cols.values())}

@router.post("/timeseries/export-excel")
def export_timeseries_excel(req: TimeSeriesExportRequest):
    years = req.years or []
    valueKeys = req.valueKeys or []
    vkUnits = req.vkUnits or {}
    entityMap = req.entityMap or {}
    entities = req.entities or sorted(entityMap.keys())

    if not years or not valueKeys or not entityMap:
        raise HTTPException(status_code=400, detail="Data deret waktu tidak lengkap")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deret Waktu"
    ws.views.sheetView[0].showGridLines = True

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    row1 = ["Rincian"]
    for y in years:
        row1.append(str(y))
        for _ in range(1, len(valueKeys)):
            row1.append("")
    ws.append(row1)

    row2 = [""]
    for _ in years:
        for vk in valueKeys:
            unit = f" ({vkUnits[vk]})" if vk in vkUnits and vkUnits[vk] else ""
            row2.append(f"{vk}{unit}")
    ws.append(row2)

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    curr_col = 2
    for y in years:
        end_col = curr_col + len(valueKeys) - 1
        if end_col > curr_col:
            ws.merge_cells(start_row=1, start_column=curr_col, end_row=1, end_column=end_col)
        curr_col = end_col + 1

    for r_idx in (1, 2):
        ws.row_dimensions[r_idx].height = 24
        for c_idx in range(1, len(row1) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            if r_idx == 1:
                cell.font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor='1E40AF')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.font = Font(name='Segoe UI', size=9, bold=True, color='1E3A8A')
                cell.fill = PatternFill('solid', fgColor='DBEAFE')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for ent in entities:
        r_vals = [ent]
        for y in years:
            y_data = entityMap.get(ent, {}).get(y, entityMap.get(ent, {}).get(str(y), {}))
            for vk in valueKeys:
                raw_v = str(y_data.get(vk, "-")).strip()
                clean_num = raw_v.replace('.', '').replace(',', '.')
                try:
                    r_vals.append(float(clean_num))
                except ValueError:
                    r_vals.append(raw_v)
        ws.append(r_vals)
        r_idx = ws.max_row
        ws.row_dimensions[r_idx].height = 20
        for c_idx, cell in enumerate(ws[r_idx], 1):
            cell.font = Font(name='Segoe UI', size=10, color='1E293B')
            cell.border = thin_border
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0.00' if isinstance(cell.value, float) and not cell.value.is_integer() else '#,##0'
            else:
                cell.alignment = Alignment(horizontal='left' if c_idx == 1 else 'center', vertical='center')

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                s = str(cell.value)
                w = sum(2 if ord(ch) > 127 else 1 for ch in s)
                if w > max_len:
                    max_len = w
        ws.column_dimensions[col_letter].width = max(min(max_len + 4, 50), 14)

    ws.freeze_panes = 'B3'
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    clean_vks = [re.sub(r'[\\/:*?"<>|]', '', vk).strip().replace(' ', '_') for vk in valueKeys]
    vk_str = "__".join([vk for vk in clean_vks if vk])[:60].rstrip('_') or "Deret_Waktu"
    sorted_years = sorted([int(y) for y in years if str(y).isdigit()])
    year_str = f"_{sorted_years[0]}-{sorted_years[-1]}" if len(sorted_years) > 1 else (f"_{sorted_years[0]}" if sorted_years else "")
    filename = f"Deret_Waktu_{vk_str}{year_str}.xlsx"
    filename = filename.encode("ascii", "ignore").decode("ascii").strip() or "Deret_Waktu.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
