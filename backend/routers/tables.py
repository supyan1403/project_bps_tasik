import os
import re
import csv
import io
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

import models
import schemas
from database import get_db
from routers.auth import require_admin, log_activity
from pipeline import deduplicate_columns, ENGLISH_ONLY_WORDS, INDO_SAFE_WORDS

router = APIRouter(prefix="/api", tags=["Tables & Editor"])

def clean_bilingual_header(header: str) -> str:
    if not header:
        return header
    header = re.sub(r'\.\d+', '', header)
    words = header.strip().split()
    if len(words) <= 1:
        return header

    while words:
        last_clean = re.sub(r'[^a-z]', '', words[-1].lower())
        if last_clean in ENGLISH_ONLY_WORDS and last_clean not in INDO_SAFE_WORDS:
            words.pop()
        else:
            break

    while len(words) > 1:
        last_word_lower = words[-1].lower()
        last_clean = re.sub(r'[^a-z]', '', last_word_lower)
        if last_clean in ENGLISH_ONLY_WORDS and last_clean not in INDO_SAFE_WORDS:
            if last_word_lower in [w.lower() for w in words[:-1]]:
                words.pop()
                continue
        break

    deduped = []
    for w in words:
        if not deduped or w.lower() != deduped[-1].lower():
            deduped.append(w)

    result = " ".join(deduped).strip()
    return result if result else header

def get_table_headers(db: Session, table) -> List[str]:
    if table and table.headers:
        return list(table.headers)
    if table:
        row = db.query(models.TableRow).filter(models.TableRow.table_id == table.id).first()
        if row and row.data:
            return list(row.data.keys())
    return []

def get_safe_windows_path(path: str) -> str:
    if not path:
        return path
    abs_p = os.path.abspath(path)
    if abs_p.startswith('\\\\?\\'):
        return abs_p
    if len(abs_p) >= 250 and os.name == 'nt':
        return '\\\\?\\' + abs_p
    return abs_p

def normalize_record_first_col(record: dict, headers: list):
    if not headers or not record:
        return
    first_key = headers[0]
    val = record.get(first_key)
    if val is not None:
        record[first_key] = str(val).strip()

@router.get("/tables/{table_id}/snippet")
def get_table_snippet(table_id: int, db: Session = Depends(get_db)):
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Tabel tidak ditemukan")
    
    doc = db.query(models.Document).filter(models.Document.id == table.document_id).first() if table.document_id else None
    try:
        orig_headers = get_table_headers(db, table) or []
        headers = [clean_bilingual_header(h) for h in orig_headers]
        units = list(table.units) if table.units else [""] * len(orig_headers)
        
        total_rows = db.query(models.TableRow).filter(models.TableRow.table_id == table_id).count()
        first_rows = db.query(models.TableRow).filter(models.TableRow.table_id == table_id).order_by(models.TableRow.sort_order.asc(), models.TableRow.id.asc()).limit(5).all()
        data_rows = [[r.data.get(h, "") for h in orig_headers] for r in first_rows]
        
        return {
            "table_id": table.id,
            "table_name": table.table_name,
            "document_name": doc.filename if doc else None,
            "document_year": doc.year if doc else None,
            "bab_num": table.bab_num if hasattr(table, "bab_num") else None,
            "total_rows": total_rows,
            "total_cols": len(headers),
            "headers": headers,
            "units": units,
            "rows": data_rows
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/tables")
def create_new_table(req: schemas.CreateTableRequest, db: Session = Depends(get_db)):
    doc = db.query(models.Document).filter(models.Document.id == req.document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Dokumen publikasi tidak ditemukan.")

    headers = [h.strip() for h in req.headers if h.strip()]
    if not headers:
        headers = ["Kecamatan", "Nilai"]

    units = list(req.units) if req.units else [""] * len(headers)
    while len(units) < len(headers):
        units.append("")

    years = list(req.years) if req.years else [str(doc.year or "")] * len(headers)
    while len(years) < len(headers):
        years.append(str(doc.year or ""))

    if units and units[0].lower() not in ["satuan", "unit"]:
        units[0] = "satuan"
    if years and years[0].lower() not in ["tahun", "year"]:
        years[0] = "tahun"

    new_table = models.ExtractedTable(
        document_id=doc.id,
        table_name=req.table_name.strip(),
        csv_path="",
        headers=headers,
        units=units,
        years=years
    )
    db.add(new_table)
    db.flush()

    initial_rows = []
    entity_col = headers[0]

    if req.rows and len(req.rows) > 0:
        for idx, r_data in enumerate(req.rows):
            row_dict = {h: r_data.get(h, "") for h in headers}
            initial_rows.append(models.TableRow(
                table_id=new_table.id,
                data=row_dict,
                is_anomaly=False,
                sort_order=idx
            ))
    elif req.custom_entities and len(req.custom_entities) > 0:
        for idx, item in enumerate(req.custom_entities):
            if not item.strip():
                continue
            row_dict = {h: "" for h in headers}
            row_dict[entity_col] = item.strip()
            initial_rows.append(models.TableRow(
                table_id=new_table.id,
                data=row_dict,
                is_anomaly=False,
                sort_order=idx
            ))
    else:
        for idx in range(10):
            row_dict = {h: "" for h in headers}
            initial_rows.append(models.TableRow(
                table_id=new_table.id,
                data=row_dict,
                is_anomaly=False,
                sort_order=idx
            ))

    db.bulk_save_objects(initial_rows)
    db.commit()
    db.refresh(new_table)
    log_activity(db, "create_table", new_table.table_name, {"table_id": new_table.id, "doc_id": doc.id})
    return {"message": "Tabel berhasil dibuat", "table_id": new_table.id, "table_name": new_table.table_name}

@router.delete("/tables/{table_id}")
def delete_single_table(table_id: int, db: Session = Depends(get_db)):
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Tabel tidak ditemukan")
    tname = table.table_name
    db.query(models.TableRow).filter(models.TableRow.table_id == table_id).delete()
    db.delete(table)
    db.commit()
    log_activity(db, "delete_table", tname or f"id={table_id}")
    return {"message": "Tabel berhasil dihapus"}

@router.get("/tables/{table_id}/excel")
@router.get("/tables/{table_id}/export_excel")
def download_table_excel(table_id: int, db: Session = Depends(get_db)):
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Tabel tidak ditemukan")
    
    headers = get_table_headers(db, table)
    if not headers:
        raise HTTPException(status_code=404, detail="Tabel tidak memiliki data kolom")
    
    units = list(table.units) if table.units else [""] * len(headers)
    years = list(table.years) if table.years else [""] * len(headers)
    while len(units) < len(headers): units.append("")
    while len(years) < len(headers): years.append("")

    all_rows = db.query(models.TableRow).filter(models.TableRow.table_id == table_id).order_by(models.TableRow.sort_order.asc(), models.TableRow.id.asc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = re.sub(r'[\\/?*\[\]:]', ' ', (table.table_name or "Tabel"))[:31]
    ws.views.sheetView[0].showGridLines = True

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    ws.append(headers)
    ws.row_dimensions[1].height = 26
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1E40AF')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    has_units = any(str(u).strip() for u in units)
    if has_units:
        ws.append(units)
        u_row = ws.max_row
        ws.row_dimensions[u_row].height = 20
        for cell in ws[u_row]:
            cell.font = Font(name='Segoe UI', size=9, italic=True, color='475569')
            cell.fill = PatternFill('solid', fgColor='F1F5F9')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    for r in all_rows:
        row_vals = []
        for h in headers:
            val_str = str(r.data.get(h, "")).strip() if isinstance(r.data, dict) else ""
            clean_num = val_str.replace('.', '').replace(',', '.')
            try:
                row_vals.append(float(clean_num))
            except ValueError:
                row_vals.append(val_str)
        ws.append(row_vals)
        r_idx = ws.max_row
        ws.row_dimensions[r_idx].height = 20
        for col_idx, cell in enumerate(ws[r_idx], 1):
            cell.font = Font(name='Segoe UI', size=10, color='1E293B')
            cell.border = thin_border
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0.00' if isinstance(cell.value, float) and not cell.value.is_integer() else '#,##0'
            else:
                cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'center', vertical='center')

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                s = str(cell.value)
                w = sum(2 if ord(ch) > 127 else 1 for ch in s)
                if w > max_len:
                    max_len = w
        ws.column_dimensions[col_letter].width = max(min(max_len + 4, 60), 14)

    ws.freeze_panes = 'A3' if has_units else 'A2'
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    clean_title = re.sub(r'[\\/:*?"<>|]', '_', (table.table_name or f"tabel_{table_id}")).strip()
    clean_title = clean_title.encode("ascii", "ignore").decode("ascii").strip() or f"tabel_{table_id}"
    filename = f"{clean_title}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@router.get("/tables/{table_id}/csv")
def download_table_csv(table_id: int, db: Session = Depends(get_db)):
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Tabel tidak ditemukan")
    
    headers = get_table_headers(db, table)
    if not headers:
        raise HTTPException(status_code=404, detail="Tabel tidak memiliki data kolom")
    
    units = table.units or [""] * len(headers)
    years = table.years or [""] * len(headers)
    all_rows = db.query(models.TableRow).filter(models.TableRow.table_id == table_id).order_by(models.TableRow.sort_order.asc(), models.TableRow.id.asc()).all()
    
    buf = io.StringIO()
    buf.write('\ufeffsep=,\r\n')
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerow(units)
    writer.writerow(years)
    for r in all_rows:
        writer.writerow([r.data.get(h, "") for h in headers])
    buf.seek(0)
    
    filename = re.sub(r'[\\/:*?"<>|]', '_', (table.table_name or f"tabel_{table_id}")) + ".csv"
    filename = filename.encode("ascii", "ignore").decode("ascii").strip() or f"tabel_{table_id}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/tables/{table_id}/csv_preview")
def preview_table_csv(table_id: int, db: Session = Depends(get_db)):
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Tabel tidak ditemukan")
    
    try:
        orig_headers = get_table_headers(db, table)
        if not orig_headers:
            return {"headers": [], "units": [], "years": [], "rows": []}
        
        units = list(table.units) if table.units else [""] * len(orig_headers)
        years = list(table.years) if table.years else [""] * len(orig_headers)
        units = ["%" if str(u).lower() in ["persen", "persentase", "percent"] else u for u in units]
        headers = [clean_bilingual_header(h) for h in orig_headers]
        
        all_rows = db.query(models.TableRow).filter(models.TableRow.table_id == table_id).order_by(models.TableRow.sort_order.asc(), models.TableRow.id.asc()).all()
        data_rows = [[r.data.get(h, "") for h in orig_headers] for r in all_rows]
        
        return {
            "table_id": table.id,
            "table_name": table.table_name,
            "headers": headers,
            "orig_headers": orig_headers,
            "units": units, 
            "years": years, 
            "rows": data_rows,
            "row_ids": [r.id for r in all_rows],
            "is_anomalies": [bool(r.is_anomaly) for r in all_rows]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class CSVRowUpdate(BaseModel):
    data: List[str]

@router.put("/tables/{table_id}/csv/row/{row_index}")
def update_csv_row(table_id: int, row_index: int, payload: CSVRowUpdate, db: Session = Depends(get_db)):
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Tabel tidak ditemukan")
    headers = get_table_headers(db, table)
    if not headers:
        raise HTTPException(status_code=400, detail="Tabel tidak memiliki kolom")
    
    rows = db.query(models.TableRow).filter(models.TableRow.table_id == table_id).order_by(models.TableRow.sort_order.asc(), models.TableRow.id.asc()).all()
    if row_index >= len(rows):
        raise HTTPException(status_code=400, detail="Row index out of range")
    
    new_data = {}
    for i, h in enumerate(headers):
        new_data[h] = payload.data[i] if i < len(payload.data) else ""
    rows[row_index].data = new_data
    db.commit()
    return {"message": "Row updated successfully"}

@router.delete("/tables/{table_id}/csv/row/{row_index}")
def delete_csv_row(table_id: int, row_index: int, db: Session = Depends(get_db)):
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Tabel tidak ditemukan")
    
    rows = db.query(models.TableRow).filter(models.TableRow.table_id == table_id).order_by(models.TableRow.sort_order.asc(), models.TableRow.id.asc()).all()
    if row_index >= len(rows):
        raise HTTPException(status_code=400, detail="Row index out of range")
    db.delete(rows[row_index])
    db.commit()
    return {"message": "Row deleted successfully"}

@router.post("/tables/{table_id}/csv/insert_row/{row_index}")
def insert_csv_row(table_id: int, row_index: int, db: Session = Depends(get_db)):
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Tabel tidak ditemukan")
    headers = get_table_headers(db, table)
    if row_index < 0:
        row_index = 0
    
    rows = db.query(models.TableRow).filter(models.TableRow.table_id == table_id).order_by(models.TableRow.sort_order.asc(), models.TableRow.id.asc()).all()
    if row_index >= len(rows):
        new_sort = (rows[-1].sort_order + 1) if rows else 0
    else:
        new_sort = rows[row_index].sort_order
        cursor = rows[row_index].sort_order
        for r in rows[row_index:]:
            r.sort_order = cursor + 1
            cursor += 1
    
    new_data = {h: "" for h in headers}
    normalize_record_first_col(new_data, headers)
    new_row = models.TableRow(table_id=table_id, data=new_data, is_anomaly=True, sort_order=new_sort)
    db.add(new_row)
    db.commit()
    return {"message": "Row inserted successfully", "insert_index": row_index}

@router.post("/tables/{table_id}/csv/row")
def add_csv_row(table_id: int, db: Session = Depends(get_db)):
    return insert_csv_row(table_id, 0, db)

class CSVColumnAdd(BaseModel):
    column_name: str
    position: Any = "end"

@router.post("/tables/{table_id}/csv/column")
def add_csv_column(table_id: int, payload: CSVColumnAdd, db: Session = Depends(get_db)):
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Tabel tidak ditemukan")
    
    headers = get_table_headers(db, table)
    col_name = payload.column_name.strip()
    if not col_name:
        raise HTTPException(status_code=400, detail="Nama kolom tidak boleh kosong")
    if col_name in headers:
        raise HTTPException(status_code=400, detail=f"Kolom '{col_name}' sudah ada")
    
    pos_config = payload.position
    insert_idx = len(headers)
    if pos_config == "start":
        insert_idx = 0
    elif pos_config == "end":
        insert_idx = len(headers)
    elif isinstance(pos_config, int):
        insert_idx = max(0, min(pos_config, len(headers)))
    elif isinstance(pos_config, dict):
        if "after_column" in pos_config:
            target = pos_config["after_column"]
            if target in headers:
                insert_idx = headers.index(target) + 1
            else:
                raise HTTPException(status_code=400, detail=f"Kolom '{target}' tidak ditemukan")
        elif "before_column" in pos_config:
            target = pos_config["before_column"]
            if target in headers:
                insert_idx = headers.index(target)
            else:
                raise HTTPException(status_code=400, detail=f"Kolom '{target}' tidak ditemukan")
    
    all_rows = db.query(models.TableRow).filter(models.TableRow.table_id == table_id).all()
    for r in all_rows:
        new_data = {}
        for i, h in enumerate(headers):
            if i == insert_idx:
                new_data[col_name] = ""
            new_data[h] = r.data.get(h, "")
        if insert_idx >= len(headers):
            new_data[col_name] = ""
        r.data = new_data
    
    table.headers = headers[:insert_idx] + [col_name] + headers[insert_idx:]
    units = list(table.units) if table.units else [""] * len(headers)
    years = list(table.years) if table.years else [""] * len(headers)
    table.units = units[:insert_idx] + [""] + units[insert_idx:]
    table.years = years[:insert_idx] + [""] + years[insert_idx:]
    db.commit()
    return {"message": "Column added successfully", "insert_index": insert_idx}

@router.delete("/tables/{table_id}/csv/column/{col_index}")
def delete_csv_column(table_id: int, col_index: int, db: Session = Depends(get_db)):
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Tabel tidak ditemukan")
    
    headers = get_table_headers(db, table)
    if col_index >= len(headers):
        raise HTTPException(status_code=400, detail="Column index out of range")
    col_name = headers[col_index]
    
    all_rows = db.query(models.TableRow).filter(models.TableRow.table_id == table_id).all()
    for r in all_rows:
        r.data = {k: v for k, v in r.data.items() if k != col_name}
    
    table.headers = [h for h in headers if h != col_name]
    if table.units:
        table.units = [u for i, u in enumerate(table.units) if i != col_index]
    if table.years:
        table.years = [y for i, y in enumerate(table.years) if i != col_index]
    db.commit()
    return {"message": "Column deleted successfully"}

class TableRenamePayload(BaseModel):
    new_name: str

@router.put("/tables/{table_id}/rename")
def rename_table(table_id: int, payload: TableRenamePayload, db: Session = Depends(get_db)):
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Table not found")
    table.table_name = payload.new_name.strip()
    db.commit()
    return {"message": "Table renamed successfully", "new_name": table.table_name}

class ColumnRenamePayload(BaseModel):
    col_index: int
    new_name: str

@router.put("/tables/{table_id}/csv/rename_column")
def rename_csv_column(table_id: int, payload: ColumnRenamePayload, db: Session = Depends(get_db)):
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Tabel tidak ditemukan")
    headers = get_table_headers(db, table)
    if not headers or payload.col_index >= len(headers):
        raise HTTPException(status_code=400, detail="Index kolom di luar batas")
    old_name = headers[payload.col_index]
    new_name = payload.new_name.strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="Nama kolom baru tidak boleh kosong")
    if new_name == old_name:
        return {"message": "Column renamed", "old_name": old_name, "new_name": new_name}
    
    all_rows = db.query(models.TableRow).filter(models.TableRow.table_id == table_id).all()
    for r in all_rows:
        if old_name in r.data:
            new_data = {}
            for k, v in r.data.items():
                new_data[new_name if k == old_name else k] = v
            r.data = new_data
    
    if table.headers:
        new_headers = list(table.headers)
        new_headers[payload.col_index] = new_name
        table.headers = new_headers
    db.commit()
    return {"message": "Column renamed", "old_name": old_name, "new_name": new_name}

class CSVSavePayload(BaseModel):
    headers: List[str]
    units: List[str]
    years: List[str]
    rows: List[List[str]]

@router.put("/tables/{table_id}/csv/save")
def save_table_csv_all(table_id: int, payload: CSVSavePayload, db: Session = Depends(get_db)):
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Tabel tidak ditemukan")
    
    if not payload.headers or all(not h.strip() for h in payload.headers):
        raise HTTPException(status_code=400, detail="Data headers tidak boleh kosong")
        
    try:
        units = list(payload.units)
        years = list(payload.years)
        if len(units) > 0:
            units[0] = "satuan"
        if len(years) > 0:
            years[0] = "tahun"
            
        for idx in range(len(units)):
            if units[idx].lower() in ["persen", "persentase", "percent"]:
                units[idx] = "%"
        headers = deduplicate_columns(payload.headers)
        
        table.headers = headers
        table.units = units
        table.years = years
        
        db.query(models.TableRow).filter(models.TableRow.table_id == table_id).delete()
        table_rows_to_add = []
        for row_arr_idx, row_arr in enumerate(payload.rows):
            record = {}
            for i, h in enumerate(headers):
                record[h] = row_arr[i] if i < len(row_arr) else ""
            normalize_record_first_col(record, headers)
            is_anomaly = False
            for val in record.values():
                str_val = str(val).strip()
                if "?" in str_val:
                    is_anomaly = True
                    break
            table_rows_to_add.append(models.TableRow(table_id=table_id, data=record, is_anomaly=is_anomaly, sort_order=row_arr_idx))
        if table_rows_to_add:
            db.bulk_save_objects(table_rows_to_add)
        db.commit()

        if table.csv_path:
            try:
                csv_path = get_safe_windows_path(table.csv_path)
                os.makedirs(os.path.dirname(csv_path), exist_ok=True)
                with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    writer.writerow(units)
                    writer.writerow(years)
                    for row_arr in payload.rows:
                        writer.writerow(row_arr)
            except Exception:
                pass

        log_activity(db, "save_table", table.table_name or f"table_id={table_id}", {"rows": len(payload.rows), "table_id": table_id})
        return {"message": "Data CSV berhasil disimpan"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/admin/tables")
def admin_get_tables(db: Session = Depends(get_db), admin: dict = Depends(require_admin)):
    tables = db.query(
        models.ExtractedTable.id, 
        models.ExtractedTable.table_name, 
        models.ExtractedTable.csv_path,
        models.Document.year,
        models.Document.filename
    ).join(models.Document, models.ExtractedTable.document_id == models.Document.id).all()
    
    result = []
    for t in tables:
        row_count = db.query(models.TableRow).filter(models.TableRow.table_id == t.id).count()
        table_name_str = t.table_name or ""
        m = re.search(r'Tabel[\s_]*(\d+)', table_name_str, re.IGNORECASE)
        result.append({
            "id": t.id,
            "table_name": table_name_str,
            "csv_path": t.csv_path,
            "year": t.year,
            "db_rows": row_count,
            "document_name": t.filename,
            "bab_num": int(m.group(1)) if m else None,
            "has_db_data": row_count > 0
        })
    return result

@router.post("/admin/clear-loaded-data")
def clear_loaded_data(db: Session = Depends(get_db), admin: dict = Depends(require_admin)):
    try:
        db.query(models.TableRow).delete()
        db.commit()
        return {"message": "All loaded table rows have been cleared successfully."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to clear data: {str(e)}")

@router.put("/admin/safe-all")
def mark_all_database_anomalies_safe(db: Session = Depends(get_db), admin: dict = Depends(require_admin)):
    count = db.query(models.TableRow).filter(models.TableRow.is_anomaly == True).count()
    db.query(models.TableRow).filter(models.TableRow.is_anomaly == True).update({"is_anomaly": False})
    db.commit()
    log_activity(db, "safe_all_anomaly", f"{count} baris ditandai aman")
    return {"message": "Semua data anomali di database berhasil ditandai aman."}

@router.get("/admin/all-data-anomalies")
def get_all_data_anomalies(db: Session = Depends(get_db), admin: dict = Depends(require_admin)):
    try:
        rows = db.query(
            models.TableRow,
            models.ExtractedTable.table_name,
            models.Document.year,
            models.Document.filename
        ).join(
            models.ExtractedTable, models.TableRow.table_id == models.ExtractedTable.id
        ).join(
            models.Document, models.ExtractedTable.document_id == models.Document.id
        ).filter(
            models.TableRow.is_anomaly == True
        ).limit(100).all()
        
        results = []
        for r, table_name, doc_year, doc_name in rows:
            m = re.search(r'Tabel[\s_]*(\d+)', table_name or "", re.IGNORECASE)
            bab_num = int(m.group(1)) if m else None
            results.append({
                "row_id": r.id,
                "table_id": r.table_id,
                "table_name": table_name,
                "document_name": doc_name,
                "document_year": doc_year,
                "bab_num": bab_num,
                "data": r.data
            })
        return {"anomalies": results}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
