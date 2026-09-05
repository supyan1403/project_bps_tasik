import io
import os
import re
import json
import zipfile
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

import models
from database import get_db
from routers.auth import log_activity, require_admin

router = APIRouter(prefix="/api/import", tags=["Excel Import"])

_BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_BPS_DATA_ROOT = os.path.join(os.path.expanduser("~"), "BPS_Data")
UPLOAD_DIR = os.path.join(_BPS_DATA_ROOT, "uploads")
try:
    os.makedirs(UPLOAD_DIR, exist_ok=True)
except Exception:
    pass

_HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_UNIT_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
_UNIT_FONT = Font(name="Calibri", italic=True, color="1E40AF", size=10)
_YEAR_FILL = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
_YEAR_FONT = Font(name="Calibri", color="475569", size=10)
_DATA_FONT = Font(name="Calibri", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def _sanitize_filename(name: str) -> str:
    if not name:
        return "template.xlsx"
    clean = str(name).strip()
    m = re.search(r'\(Hal[_\s]+([\d,_\s]+)\)', clean, re.IGNORECASE)
    if m:
        pages = re.findall(r'\d+', m.group(1))
        if len(pages) > 2:
            new_hal = f"(Hal_{pages[0]}-{pages[-1]})"
            clean = clean[:m.start()] + new_hal + clean[m.end():]
    clean = re.sub(r'[\\/:*?"<>|]', '', clean)
    clean = clean.replace(' ', '_')
    if len(clean) > 120:
        clean = clean[:120].rstrip('_- ')
    return f"{clean}.xlsx"


def _build_template_workbook(
    table_name: str,
    headers: list,
    units: list,
    years: list,
    pub_year: Optional[int] = None,
    col_year: Optional[int] = None,
) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # --- Sheet "Info" ---
    ws_info = wb.active
    ws_info.title = "Info"
    ws_info.column_dimensions['A'].width = 22
    ws_info.column_dimensions['B'].width = 60

    info_rows = [
        ("Nama Tabel", table_name or ""),
        ("Tahun Publikasi", pub_year or ""),
        ("Tahun Data", col_year or ""),
        ("Jumlah Kolom", len(headers) or ""),
        ("Kolom (JSON)", json.dumps(headers or [], ensure_ascii=False)),
        ("Satuan (JSON)", json.dumps(units or [], ensure_ascii=False)),
        ("Tahun (JSON)", json.dumps(years or [], ensure_ascii=False)),
    ]
    for r_idx, (label, val) in enumerate(info_rows, start=1):
        cell_a = ws_info.cell(row=r_idx, column=1, value=label)
        cell_a.font = Font(bold=True, size=10)
        cell_b = ws_info.cell(row=r_idx, column=2, value=str(val))
        cell_b.font = Font(size=10)

    # --- Sheet "Data" ---
    ws_data = wb.create_sheet("Data")

    # Row 1: Headers
    for c_idx, h in enumerate(headers or ["Kolom 1"], start=1):
        cell = ws_data.cell(row=1, column=c_idx, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    ws_data.row_dimensions[1].height = 22

    # Row 2: Units
    for c_idx, u in enumerate(units or [""] * len(headers or [""]), start=1):
        cell = ws_data.cell(row=2, column=c_idx, value=u if u else "")
        cell.fill = _UNIT_FILL
        cell.font = _UNIT_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER
    ws_data.row_dimensions[2].height = 18

    # Row 3: Years
    for c_idx, y in enumerate(years or [""] * len(headers or [""]), start=1):
        cell = ws_data.cell(row=3, column=c_idx, value=y if y else "")
        cell.fill = _YEAR_FILL
        cell.font = _YEAR_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER
    ws_data.row_dimensions[3].height = 18

    # Row 4-13: 10 empty data rows
    for r_idx in range(4, 14):
        for c_idx in range(1, len(headers or [""]) + 1):
            cell = ws_data.cell(row=r_idx, column=c_idx, value="")
            cell.font = _DATA_FONT
            cell.border = _THIN_BORDER

    # Auto-fit column widths (approximate)
    for c_idx in range(1, len(headers or [""]) + 1):
        header_len = len(str((headers or [""])[c_idx - 1] or ""))
        ws_data.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = max(14, min(header_len + 4, 35))

    return wb


@router.get("/template")
def download_single_template(
    table_id: int,
    pub_year: Optional[int] = None,
    col_year: Optional[int] = None,
    db: Session = Depends(get_db),
):
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Tabel tidak ditemukan.")

    headers = table.headers or []
    units = table.units or []
    years = table.years or []

    if not pub_year and table.document:
        pub_year = table.document.year
    if not col_year and pub_year:
        col_year = pub_year - 1

    wb = _build_template_workbook(table.table_name, headers, units, years, pub_year, col_year)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = _sanitize_filename(table.table_name)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/template/zip")
def download_bab_zip(
    doc_id: int,
    bab_num: int,
    pub_year: Optional[int] = None,
    col_year: Optional[int] = None,
    db: Session = Depends(get_db),
):
    doc = db.query(models.Document).filter(models.Document.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Publikasi tidak ditemukan.")

    tables = (
        db.query(models.ExtractedTable)
        .filter(models.ExtractedTable.document_id == doc_id)
        .all()
    )

    if not tables:
        raise HTTPException(status_code=404, detail="Tidak ada tabel ditemukan untuk publikasi ini.")

    filtered = []
    for t in tables:
        match = re.search(r'Tabel[\s_]*(\d+)', t.table_name or '', re.IGNORECASE)
        t_bab = int(match.group(1)) if match else 999
        if t_bab == bab_num:
            filtered.append(t)

    if not filtered:
        raise HTTPException(status_code=404, detail=f"Tidak ada tabel ditemukan untuk Bab {bab_num}.")

    if not pub_year:
        pub_year = doc.year
    if not col_year and pub_year:
        col_year = pub_year - 1

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for t in filtered:
            wb = _build_template_workbook(
                t.table_name, t.headers or [], t.units or [], t.years or [],
                pub_year, col_year,
            )
            tpl_buf = io.BytesIO()
            wb.save(tpl_buf)
            tpl_buf.seek(0)
            fname = _sanitize_filename(t.table_name)
            zf.writestr(fname, tpl_buf.read())

    zip_buf.seek(0)
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="templates_Bab{bab_num}.zip"'},
    )


@router.post("/excel")
def import_excel_files(
    files: list[UploadFile] = File(...),
    year: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    admin: dict = Depends(require_admin)
):
    if not files:
        raise HTTPException(status_code=400, detail="Tidak ada file yang diunggah.")

    imported_docs = []
    total_tables = 0
    total_rows = 0
    errors = []

    for upload_file in files:
        try:
            content = upload_file.file.read()
            if not content:
                errors.append(f"{upload_file.filename}: file kosong")
                continue

            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

            # --- Parse Sheet "Info" ---
            table_name = ""
            headers = []
            units = []
            years = []
            doc_year = year

            if "Info" in wb.sheetnames:
                ws_info = wb["Info"]
                info_map = {}
                for row in ws_info.iter_rows(min_row=1, max_col=2, values_only=True):
                    if row[0]:
                        info_map[str(row[0]).strip()] = row[1]

                table_name = str(info_map.get("Nama Tabel", "")).strip()
                raw_kolom = info_map.get("Kolom (JSON)", "[]")
                raw_satuan = info_map.get("Satuan (JSON)", "[]")
                raw_tahun = info_map.get("Tahun (JSON)", "[]")

                try:
                    headers = json.loads(raw_kolom) if isinstance(raw_kolom, str) else raw_kolom or []
                except Exception:
                    headers = []
                try:
                    units = json.loads(raw_satuan) if isinstance(raw_satuan, str) else raw_satuan or []
                except Exception:
                    units = []
                try:
                    years = json.loads(raw_tahun) if isinstance(raw_tahun, str) else raw_tahun or []
                except Exception:
                    years = []

                if not doc_year:
                    py = info_map.get("Tahun Publikasi")
                    if py:
                        try:
                            doc_year = int(py)
                        except Exception:
                            pass

            if not table_name:
                table_name = upload_file.filename or "Tabel Import"

            # Fallback: infer headers from Data sheet if Info sheet is empty
            if not headers and "Data" in wb.sheetnames:
                ws_data = wb["Data"]
                headers = [str(c.value or "").strip() for c in next(ws_data.iter_rows(min_row=1, max_row=1))]
                headers = [h if h else f"Kolom {i+1}" for i, h in enumerate(headers)]
                if len(units) < len(headers):
                    units = [str(c.value or "").strip() for c in next(ws_data.iter_rows(min_row=2, max_row=2))]
                    units = units + [""] * (len(headers) - len(units))
                if len(years) < len(headers):
                    years_raw = [str(c.value or "").strip() for c in next(ws_data.iter_rows(min_row=3, max_row=3))]
                    years = years_raw + [""] * (len(headers) - len(years_raw))

            # --- Create Document ---
            doc = models.Document(
                filename=upload_file.filename or f"import_{table_name}.xlsx",
                year=doc_year,
                status="ready",
            )
            db.add(doc)
            db.flush()

            # --- Create ExtractedTable ---
            ext_table = models.ExtractedTable(
                document_id=doc.id,
                table_name=table_name,
                headers=headers,
                units=units,
                years=years,
                csv_path="",
            )
            db.add(ext_table)
            db.flush()
            total_tables += 1

            # --- Parse Sheet "Data" ---
            if "Data" in wb.sheetnames:
                ws_data = wb["Data"]
                row_count = 0
                for row_idx, row in enumerate(ws_data.iter_rows(min_row=4, values_only=True), start=0):
                    # Skip entirely empty rows
                    vals = [c for c in row if c is not None and str(c).strip() != ""]
                    if not vals:
                        continue

                    row_dict = {}
                    for c_idx, h in enumerate(headers):
                        val = row[c_idx] if c_idx < len(row) else None
                        row_dict[h] = str(val).strip() if val is not None else ""

                    # Skip rows where the first column (label) is empty
                    if not row_dict.get(headers[0], "").strip():
                        continue

                    tr = models.TableRow(
                        table_id=ext_table.id,
                        data=row_dict,
                        sort_order=row_count,
                    )
                    db.add(tr)
                    row_count += 1
                    total_rows += 1

                ext_table.csv_path = f"imported/{doc.id}/{ext_table.id}.csv"

            imported_docs.append({
                "doc_id": doc.id,
                "filename": doc.filename,
                "tables": 1,
            })

        except Exception as e:
            errors.append(f"{upload_file.filename}: {str(e)}")

    db.commit()

    message_parts = []
    if imported_docs:
        message_parts.append(f"Berhasil mengimpor {len(imported_docs)} file ({total_tables} tabel, {total_rows} baris data).")
    if errors:
        message_parts.append(f"Error pada {len(errors)} file: {'; '.join(errors[:3])}")

    first_doc_id = imported_docs[0]["doc_id"] if imported_docs else 0

    log_activity(
        db, "import_excel",
        f"Import {len(imported_docs)} file Excel",
        {"total_tables": total_tables, "total_rows": total_rows, "errors": len(errors)},
    )

    return {
        "message": " ".join(message_parts) or "Tidak ada data yang diimpor.",
        "document_id": first_doc_id,
        "total_tables": total_tables,
        "total_rows": total_rows,
    }
