import csv
import io
import json
import os
import re
import secrets
import subprocess
import sys
import threading
import time
import unicodedata
import zipfile
from contextlib import asynccontextmanager
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Union, Optional
from pydantic import BaseModel
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form, BackgroundTasks, Request, Body, Response, Cookie
from fastapi.responses import StreamingResponse, JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from sqlalchemy import inspect, text

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import models
import schemas
from database import engine, get_db, SessionLocal
# ... [imports and utility functions] ...
from pipeline import detect_and_clean_metadata, deduplicate_columns, ENGLISH_ONLY_WORDS, INDO_SAFE_WORDS, parse_indonesian_number

# Trigger reload to load updated INDO_SAFE_WORDS from pipeline_utils
def clean_bilingual_header(header: str) -> str:
    """
    Bersihkan nama kolom bilingual duplikat, contoh:
    "Irigasi Irrigation"           -> "Irigasi"
    "Non Irigasi Non Irrigation"   -> "Non Irigasi"
    "Lahan Sawah Paddy Field"      -> "Lahan Sawah"
    
    Strategi:
    1. Hapus kata-kata bahasa Inggris dari belakang.
    2. Hapus kata-kata trailing yang sudah muncul di bagian awal (duplikat parsial).
    3. Hapus konsekutif duplikat kata (misal "Sawah Sawah" -> "Sawah").
    """
    if not header:
        return header

    # Remove pandas-induced duplicate suffixes (e.g., .1, .2)
    header = re.sub(r'\.\d+', '', header)

    words = header.strip().split()
    if len(words) <= 1:
        return header

    # Langkah 1: Hapus kata Inggris dari belakang
    while words:
        last_clean = re.sub(r'[^a-z]', '', words[-1].lower())
        if last_clean in ENGLISH_ONLY_WORDS and last_clean not in INDO_SAFE_WORDS:
            words.pop()
        else:
            break

    # Langkah 2: Hapus kata trailing yang sudah ada di bagian awal (menangani "Non Irigasi Non")
    # Cek apakah kata terakhir sudah muncul sebelumnya di kata-kata sebelumnya
    while len(words) > 1:
        last_word_lower = words[-1].lower()
        last_clean = re.sub(r'[^a-z]', '', last_word_lower)
        if last_clean in ENGLISH_ONLY_WORDS and last_clean not in INDO_SAFE_WORDS:
            # Cek apakah kata terakhir juga ada di bagian sebelumnya
            if last_word_lower in [w.lower() for w in words[:-1]]:
                words.pop()
                continue
        break

    # Langkah 3: Hapus kata duplikat berturut-turut (misal "Sawah Sawah" -> "Sawah")
    deduped = []
    for w in words:
        if not deduped or w.lower() != deduped[-1].lower():
            deduped.append(w)

    result = " ".join(deduped).strip()
    return result if result else header


# =====================================================================
# HELPER BACA METADATA TABEL LANGSUNG DARI DATABASE
# (alur data tidak bergantung pada file CSV lokal)
# =====================================================================
def get_table_headers(db: Session, table) -> List[str]:
    """Ambil daftar nama kolom tabel. Prioritas: metadata tersimpan -> kunci data baris pertama."""
    if table and table.headers:
        return list(table.headers)
    if table:
        row = db.query(models.TableRow).filter(models.TableRow.table_id == table.id).first()
        if row and row.data:
            return list(row.data.keys())
    return []


def extract_timeseries_year(year_val, default_year=None):
    """
    Mengekstrak tahun awal untuk keperluan Analisis Deret Waktu (Time-Series Analysis)
    dari berbagai format string tahun ajaran / rentang tahun / angka:
    - '2020/2021' -> 2020
    - '2021/2022' -> 2021
    - '2020-2021' -> 2020
    - '2020/21'   -> 2020
    - '2020'      -> 2020
    - 2021        -> 2021
    """
    if year_val is None or year_val == '':
        return default_year
    s = str(year_val).strip()
    m = re.search(r'\b(19\d{2}|20\d{2})\b', s)
    if m:
        return int(m.group(1))
    return default_year


def get_column_years_from_db(db: Session, table, doc_year: int) -> dict:
    """Versi DB dari get_column_years_from_csv. Menentukan tahun per kolom
    dari metadata 'years' tersimpan atau dari isi data (untuk tabel vertikal).
    Nama kolom dikembalikan dalam bentuk bersih (konvensi master kolom,
    suffix deduplikasi seperti '.1' dibuang)."""
    headers = get_table_headers(db, table)
    if not headers:
        return {}
    years = table.years or []
    col_years = {}

    first_lower = headers[0].lower() if headers else ""
    is_vertical = "tahun" in first_lower or "year" in first_lower

    if is_vertical:
        rows = db.query(models.TableRow).filter(models.TableRow.table_id == table.id).all()
        years_in_col0 = set()
        for r in rows:
            val = str(r.data.get(headers[0], "")).strip()
            y_extracted = extract_timeseries_year(val)
            if y_extracted:
                years_in_col0.add(y_extracted)
        for idx, h in enumerate(headers):
            if idx == 0:
                continue
            if not any(char.isalpha() for char in h):
                continue
            col_years.setdefault(_clean_header_for_master(h), set()).update(years_in_col0)
        col_years = {k: sorted(list(v)) for k, v in col_years.items()}
    else:
        for idx, h in enumerate(headers):
            if idx == 0:
                continue
            if not any(char.isalpha() for char in h):
                continue
            yr_val = years[idx] if idx < len(years) else ""
            col_year = extract_timeseries_year(yr_val, default_year=doc_year - 1)
            col_years.setdefault(_clean_header_for_master(h), set()).add(col_year)
        col_years = {k: sorted(list(v)) for k, v in col_years.items()}

    return col_years



def get_column_years_with_position(db: Session, table, doc_year: int) -> dict:
    """Serupa dengan get_column_years_from_db, namun juga mengembalikan posisi
    kolom (idx) untuk setiap nama kolom. Return: {cleaned_name: {"years": set, "pos": idx}}."""
    headers = get_table_headers(db, table)
    if not headers:
        return {}
    years = table.years or []
    result = {}
    first_lower = headers[0].lower() if headers else ""
    is_vertical = "tahun" in first_lower or "year" in first_lower

    if is_vertical:
        rows = db.query(models.TableRow).filter(models.TableRow.table_id == table.id).all()
        years_in_col0 = set()
        for r in rows:
            val = str(r.data.get(headers[0], "")).strip()
            y_extracted = extract_timeseries_year(val)
            if y_extracted:
                years_in_col0.add(y_extracted)
        for idx, h in enumerate(headers):
            if idx == 0:
                continue
            if not any(char.isalpha() for char in h):
                continue
            name = _clean_header_for_master(h)
            result.setdefault(name, {"years": set(), "pos": idx})
            result[name]["years"].update(years_in_col0)
    else:
        for idx, h in enumerate(headers):
            if idx == 0:
                continue
            if not any(char.isalpha() for char in h):
                continue
            name = _clean_header_for_master(h)
            yr_val = years[idx] if idx < len(years) else ""
            col_year = extract_timeseries_year(yr_val, default_year=doc_year - 1)
            result.setdefault(name, {"years": set(), "pos": idx})
            result[name]["years"].add(col_year)

    return result



def get_safe_windows_path(path: str) -> str:
    if os.name == 'nt':
        abs_path = os.path.abspath(path)
        if not abs_path.startswith("\\\\?\\"):
            return "\\\\?\\" + abs_path.replace("/", "\\")
    return path


try:
    models.Base.metadata.create_all(bind=engine)
except Exception as e:
    print(f"[db] Peringatan create_all dilewati atau izin terbatas: {e}")

# ===== MIGRASI KOLOM BARU (idempoten) =====
def migrate_db_columns():
    """Tambahkan kolom baru ke tabel yang sudah ada tanpa menghapus data.
    Diperlukan karena create_all tidak menambah kolom ke tabel SQLite lama
    (headers/units/years di extracted_tables, sort_order di table_rows)."""
    inspector = inspect(engine)
    with engine.begin() as conn:
        existing_et = {c["name"] for c in inspector.get_columns("extracted_tables")}
        for col, coltype in [("headers", "JSON"), ("units", "JSON"), ("years", "JSON")]:
            if col not in existing_et:
                conn.execute(text(f"ALTER TABLE extracted_tables ADD COLUMN {col} {coltype}"))
        existing_tr = {c["name"] for c in inspector.get_columns("table_rows")}
        if "sort_order" not in existing_tr:
            conn.execute(text("ALTER TABLE table_rows ADD COLUMN sort_order INTEGER"))

try:
    migrate_db_columns()
    print("[migrate] Kolom database diverifikasi/diperbarui.")
except Exception as e:
    print(f"[migrate] Peringatan: gagal memigrasi kolom database: {e}")

@asynccontextmanager
async def lifespan(app: FastAPI):
    _db = next(get_db())
    _clean_expired_sessions(_db)
    _db.close()
    reset_stuck_extractions()
    yield

app = FastAPI(title="BPS Extraction Dashboard API", lifespan=lifespan)

def reset_stuck_extractions():
    db = next(get_db())
    try:
        stuck_docs = db.query(models.Document).filter(models.Document.status.like("extracting%")).all()
        for doc in stuck_docs:
            doc.status = "ready"
        db.commit()
        print(f"Reset {len(stuck_docs)} stuck document extraction status(es) to ready.")
    except Exception as e:
        print(f"Gagal me-reset status ekstraksi terhenti: {e}")

_PROD_DOMAIN = os.environ.get("SIPEDAS_DOMAIN", "")
if _PROD_DOMAIN:
    _CORS_ORIGINS = [f"https://{_PROD_DOMAIN}", f"http://{_PROD_DOMAIN}"]
else:
    _CORS_ORIGINS = ["http://127.0.0.1:8000", "http://localhost:8000"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=_CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
# Kompresi GZip otomatis untuk respon JSON & berkas statis > 1KB (menghemat transfer 70-85%)
app.add_middleware(GZipMiddleware, minimum_size=1000)

# Caching cerdas untuk aset statis (CSS, JS, Fonts, Gambar)
@app.middleware("http")
async def add_cache_control_header(request: Request, call_next):
    response = await call_next(request)
    if request.url.path.startswith("/static/"):
        response.headers["Cache-Control"] = "public, max-age=0, must-revalidate"
    return response

# =====================================================================
# MAINTENANCE MODE — database-backed via SystemConfig table
# Key: "maintenance_mode" = "1"|"0", "maintenance_end" = ISO timestamp
# Auto-disable saat maintenance_end lewat.
# Admin yang sudah login tetap bisa akses selama maintenance.
# =====================================================================

_maintenance_cache = {"active": False, "end": "", "ts": 0}
_CACHE_TTL = 5  # detik

def _is_maintenance():
    now = time.time()
    if now - _maintenance_cache["ts"] < _CACHE_TTL:
        return _maintenance_cache["active"]
    try:
        db = SessionLocal()
        row = db.query(models.SystemConfig).filter(models.SystemConfig.key == "maintenance_mode").first()
        val = row.value.strip() if row else "0"
        end_row = db.query(models.SystemConfig).filter(models.SystemConfig.key == "maintenance_end").first()
        end_val = end_row.value.strip() if end_row else ""
        db.close()
        # Auto-disable jika maintenance_end sudah lewat
        if val == "1" and end_val:
            from datetime import datetime, timedelta
            try:
                clean_val = end_val.replace('Z', '').split('+')[0].split('.')[0]
                end_dt = datetime.fromisoformat(clean_val)
                # Jika punya Z/offset → UTC. Jika tidak → asumsi local time
                is_utc = 'Z' in end_val or '+' in end_val or (len(end_val) > 19 and end_val[19] in '+-')
                now_dt = datetime.utcnow() if is_utc else datetime.now()
                if now_dt >= end_dt:
                    val = "0"
                    _update_maintenance_db("0", "")
            except Exception:
                pass
        _maintenance_cache["active"] = (val == "1")
        _maintenance_cache["end"] = end_val
        _maintenance_cache["ts"] = now
    except Exception:
        pass
    return _maintenance_cache["active"]

def _maintenance_end():
    if _maintenance_cache["ts"] and (time.time() - _maintenance_cache["ts"] < _CACHE_TTL):
        return _maintenance_cache["end"]
    _is_maintenance()  # refresh cache
    return _maintenance_cache["end"]

def _update_maintenance_db(mode, end_time=""):
    try:
        db = SessionLocal()
        for k, v in [("maintenance_mode", mode), ("maintenance_end", end_time)]:
            row = db.query(models.SystemConfig).filter(models.SystemConfig.key == k).first()
            if row:
                row.value = v
            else:
                db.add(models.SystemConfig(key=k, value=v))
        db.commit()
        db.close()
    except Exception:
        pass
        pass
    _maintenance_cache["ts"] = 0

@app.middleware("http")
async def maintenance_middleware(request: Request, call_next):
    if _is_maintenance():
        path = request.url.path
        # Static files tetap jalan
        if path.startswith("/static/") or path == "/favicon.ico":
            return await call_next(request)

        # Login page & login API harus tetap bisa diakses
        # Supaya admin bisa login → lalu bypass maintenance
        if path == "/login" or path.startswith("/api/auth/login"):
            return await call_next(request)

        # Force maintenance dari cross-tab sync: skip admin bypass
        force_maintenance = '_force_maintenance=1' in str(request.url.query)

        # Cek apakah user adalah admin yang sudah login (skip jika force)
        if not force_maintenance:
            session_id = request.cookies.get("sipedas_session")
            if session_id:
                db = next(get_db())
                try:
                    sess = db.query(models.UserSession).filter(
                        models.UserSession.id == session_id,
                        models.UserSession.role == "admin"
                    ).first()
                    if sess:
                        return await call_next(request)
                finally:
                    db.close()

        # Non-admin → maintenance page
        if path.startswith("/api/"):
            return JSONResponse(status_code=503, content={"detail": "Sistem sedang dalam pemeliharaan. Silakan coba lagi nanti."})
        return templates.TemplateResponse(
            request=request, name="maintenance.html", status_code=503,
            context={"maintenance_end": _maintenance_end()}
        )
    return await call_next(request)

# =====================================================================
# GLOBAL ERROR HANDLERS — 404, 500, dan error tak terduga
# =====================================================================
@app.exception_handler(404)
async def not_found_handler(request: Request, exc):
    path = request.url.path
    if path.startswith("/api/"):
        return JSONResponse(status_code=404, content={"detail": "Endpoint tidak ditemukan."})
    return templates.TemplateResponse(request=request, name="404.html", status_code=404)

@app.exception_handler(500)
async def internal_error_handler(request: Request, exc):
    path = request.url.path
    if path.startswith("/api/"):
        return JSONResponse(status_code=500, content={"detail": "Terjadi kesalahan internal server. Silakan coba lagi."})
    return templates.TemplateResponse(request=request, name="500.html", status_code=500)

@app.exception_handler(Exception)
async def general_exception_handler(request: Request, exc):
    print(f"[ERROR] Unhandled exception at {request.url.path}: {exc}")
    path = request.url.path
    if path.startswith("/api/"):
        return JSONResponse(status_code=500, content={"detail": "Terjadi kesalahan tidak terduga. Silakan coba lagi."})
    return templates.TemplateResponse(request=request, name="500.html", status_code=500)

# =====================================================================
# AUTH ROUTER — Session-based admin authentication
# =====================================================================
from routers.auth import router as auth_router, require_admin, get_current_role, log_activity, _clean_expired_sessions
app.include_router(auth_router)

_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(_BASE_DIR, "static")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
templates = Jinja2Templates(directory=os.path.join(_BASE_DIR, "templates"))

# =====================================================================
# DIREKTORI PENYIMPANAN
# CSV hasil ekstraksi dan PDF upload disimpan di luar folder project
# agar tidak mengotori direktori kode sumber.
# Lokasi: ~/BPS_Data/ (Linux/Mac) atau C:\Users\[user]\BPS_Data\ (Windows)
# =====================================================================
_BPS_DATA_ROOT = os.path.join(os.path.expanduser("~"), "BPS_Data")
UPLOAD_DIR = os.path.join(_BPS_DATA_ROOT, "uploads")
EXTRACT_DIR = os.path.join(_BPS_DATA_ROOT, "hasil_ekstraksi_web")

# Backup database .sql disimpan di folder backups/ dalam project (di luar hasil_ekstraksi_web)
BACKUP_DIR = os.path.abspath(os.path.join(_BASE_DIR, "..", "backups"))

try:
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    os.makedirs(EXTRACT_DIR, exist_ok=True)
    os.makedirs(BACKUP_DIR, exist_ok=True)
except Exception as _e:
    pass

@app.get("/")
def read_root(request: Request):
    response = templates.TemplateResponse(request=request, name="index.html", context={})
    response.headers["Cache-Control"] = "no-store"
    return response

@app.get("/login")
def login_page(request: Request):
    response = templates.TemplateResponse(request=request, name="index.html", context={})
    response.headers["Cache-Control"] = "no-store"
    return response

@app.get("/favicon.ico", include_in_schema=False)
def get_favicon():
    favicon_path = os.path.join(STATIC_DIR, "logo_sipedas.png")
    if os.path.exists(favicon_path):
        return FileResponse(favicon_path, media_type="image/png")
    return Response(status_code=204)

@app.get("/api/stats")
def get_dashboard_stats(db: Session = Depends(get_db)):
    total_docs = db.query(models.Document).count()
    total_tables = db.query(models.ExtractedTable).count()
    total_rows = db.query(models.TableRow).count()
    total_anomalies = db.query(models.TableRow).filter(models.TableRow.is_anomaly == True).count()
    
    # Hitung total titik nilai data statistik (sel data terisi angka/nilai, exclude kolom label dan tanda strip '-')
    LABEL_KEYS = {'Kecamatan', 'No / Uraian', 'Uraian', 'No', 'Desa', 'Kelurahan', 'Desa/Kelurahan'}
    EMPTY_MARKERS = {'', '-', '--', '...', 'nan', 'none', 'null'}
    total_data_points = 0
    try:
        rows = db.query(models.TableRow.data).all()
        total_data_points = sum(
            len([v for k, v in (r.data or {}).items()
                 if k not in LABEL_KEYS and v is not None and str(v).strip().lower() not in EMPTY_MARKERS])
            for r in rows if r.data
        )
    except Exception:
        total_data_points = 0

    avg_rows_per_table = round(total_rows / max(total_tables, 1), 1)
    avg_points_per_table = round(total_data_points / max(total_tables, 1), 1)
    
    # Hitung rentang tahun publikasi
    doc_years = [d.year for d in db.query(models.Document.year).all() if d.year]
    year_range = f"{min(doc_years)} - {max(doc_years)}" if doc_years else "2023 - 2026"
    
    # Hitung total bab BPS unik yang terdeteksi
    all_table_names = db.query(models.ExtractedTable.table_name).all()
    detected_babs = set()
    for (tname,) in all_table_names:
        m = re.search(r'Tabel\s+(\d+)\.', tname or '')
        if m:
            detected_babs.add(int(m.group(1)))
    total_babs = len(detected_babs) if detected_babs else 13

    return {
        "total_docs": total_docs,
        "total_tables": total_tables,
        "total_rows": total_rows,
        "total_data_points": total_data_points,
        "total_anomalies": total_anomalies,
        "avg_rows_per_table": avg_rows_per_table,
        "avg_points_per_table": avg_points_per_table,
        "total_babs": total_babs,
        "year_range": year_range
    }

# =====================================================================
# ADMIN & BACKUP ROUTER
# =====================================================================
from routers.admin import router as admin_router, backup_database, cleanup_old_backups, restore_database
app.include_router(admin_router)
# =====================================================================
# DASHBOARD STATS ROUTER
# =====================================================================
from routers.stats import router as stats_router
app.include_router(stats_router)
# =====================================================================
# TABLES & EDITOR ROUTER
# =====================================================================
from routers.tables import router as tables_router, get_table_headers, clean_bilingual_header, normalize_record_first_col
app.include_router(tables_router)

# =====================================================================
# DOCUMENTS & EXTRACTION ROUTER
# =====================================================================
from routers.documents import router as documents_router
app.include_router(documents_router)

# =====================================================================
# TIME SERIES ANALYSIS ROUTER
# =====================================================================
from routers.timeseries import (
    router as timeseries_router,
    extract_timeseries_year,
    get_column_years_from_db,
    get_column_years_with_position,
    get_clean_chapter_name,
    get_clean_table_name,
    normalize_entity_name,
    normalize_entity_key,
    _extract_year_range_from_table_name,
    _get_unit_multiplier,
    _format_scaled_indo_number,
    _normalize_base_unit,
    _normalize_indo_number,
    _extract_unit_from_table_name,
    _infer_unit_from_indicator,
    _infer_unit_from_headers_and_table,
    _build_vk_units,
    _classify_entity_type,
    check_cell_format_anomaly,
    detect_timeseries_anomalies,
    _dedup_timeseries_results
)
app.include_router(timeseries_router)

# =====================================================================
# ANOMALY DETECTION ROUTER
# =====================================================================
from routers.anomaly import (
    router as anomaly_router,
    _get_ttl_cache,
    _set_ttl_cache,
    _clear_ttl_cache,
    _load_ts_safe,
    _save_ts_safe,
    _load_master_dict,
    _save_master_dict,
    _load_dismissed,
    _save_dismissed,
    _load_master_columns,
    get_timeseries_anomalies
)
app.include_router(anomaly_router)

# =====================================================================
# MASTER COLUMNS & DICTIONARIES ROUTER
# =====================================================================
from routers.master_data import (
    router as master_data_router,
    _get_master_col_unit_map,
    _clean_header_for_master,
    suggest_master_columns
)
app.include_router(master_data_router)

# =====================================================================
# EXCEL IMPORT ROUTER
# =====================================================================
from routers.import_excel import router as import_excel_router
app.include_router(import_excel_router)

STANDARD_TASIK_KECAMATAN = [
    "Cipatujah", "Karangnunggal", "Cikalong", "Pancatengah", "Cikatomas",
    "Cibalong", "Parungponteng", "Bantarkalong", "Bojongasih", "Culamega",
    "Bojonggambir", "Sodonghilir", "Taraju", "Salawu", "Puspahiang",
    "Tanjungjaya", "Sukaraja", "Salopa", "Jatiwaras", "Cineam",
    "Karangjaya", "Manonjaya", "Gunungtanjung", "Singaparna", "Mangunreja",
    "Sukarame", "Cigalontang", "Leuwisari", "Padakembang", "Sariwangi",
    "Sukaratu", "Cisayong", "Sukahening", "Rajapolah", "Jamanis",
    "Ciawi", "Kadipaten", "Pagerageung", "Sukaresik", "Kabupaten Tasikmalaya"
]

STANDARD_BULAN = [
    "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember", "Tahunan / Total"
]

STANDARD_LAPANGAN_USAHA_PDRB = [
    "A. Pertanian, Kehutanan, dan Perikanan",
    "B. Pertambangan dan Penggalian",
    "C. Industri Pengolahan",
    "D. Pengadaan Listrik dan Gas",
    "E. Pengadaan Air, Pengelolaan Sampah, Limbah dan Daur Ulang",
    "F. Konstruksi",
    "G. Perdagangan Besar dan Eceran; Reparasi Mobil dan Sepeda Motor",
    "H. Transportasi dan Pergudangan",
    "I. Penyediaan Akomodasi dan Makan Minum",
    "J. Informasi dan Komunikasi",
    "K. Jasa Keuangan dan Asuransi",
    "L. Real Estat",
    "M,N. Jasa Perusahaan",
    "O. Administrasi Pemerintahan, Pertahanan dan Jaminan Sosial Wajib",
    "P. Jasa Pendidikan",
    "Q. Jasa Kesehatan dan Kegiatan Sosial",
    "R,S,T,U. Jasa Lainnya",
    "Produk Domestik Regional Bruto (PDRB)"
]

STANDARD_KOMODITAS_PERTANIAN = [
    "Padi Sawah", "Padi Ladang", "Jagung", "Kedelai", "Kacang Tanah",
    "Ubi Kayu", "Ubi Jalar", "Cabai Rawit", "Cabai Merah", "Bawang Merah", "Tomat", "Pisang", "Jumlah"
]

STANDARD_JENIS_TERNAK = [
    "Sapi Potong", "Sapi Perah", "Kerbau", "Kuda", "Kambing",
    "Domba", "Ayam Buras", "Ayam Ras Pedaging", "Ayam Ras Petelur", "Itik / Bebek", "Jumlah"
]

STANDARD_KELOMPOK_UMUR = [
    "0-4", "5-9", "10-14", "15-19", "20-24", "25-29", "30-34",
    "35-39", "40-44", "45-49", "50-54", "55-59", "60-64", "65-69", "70-74", "75+", "Jumlah"
]

def parse_csv_for_db(safe_path: str) -> tuple:
    raw_rows = []
    with open(safe_path, 'r', encoding='utf-8', errors='replace') as f:
        reader = csv.reader(f)
        for row in reader:
            raw_rows.append(row)
            
    if not raw_rows:
        return [], [], [], []
        
    headers = raw_rows[0]
    # Deduplikasi nama kolom yang sama agar tidak ada data yang hilang diam-diam.
    # "A,A" -> "A","A.1"; "A.1,A" -> tetap unik. (konsisten dgn ekstraktor)
    seen = {}
    for i, h in enumerate(headers):
        key = str(h)
        cnt = seen.get(key, 0) + 1
        seen[key] = cnt
        if cnt > 1:
            headers[i] = f"{h}.{cnt - 1}"
    units = [""] * len(headers)
    years = [""] * len(headers)
    has_metadata = False
    if len(raw_rows) >= 3:
        col0_row1 = str(raw_rows[1][0]).strip().lower() if len(raw_rows[1]) > 0 else ""
        col0_row2 = str(raw_rows[2][0]).strip().lower() if len(raw_rows[2]) > 0 else ""
        if col0_row1 == "satuan" or col0_row2 == "tahun":
            has_metadata = True
            units = raw_rows[1]
            years = raw_rows[2]
        else:
            row1 = raw_rows[1]
            row2 = raw_rows[2]
            row1_col0_empty = str(row1[0]).strip() == "" if row1 else True
            row1_has_content = any(str(v).strip() for v in row1[1:]) if len(row1) > 1 else False
            row2_col0_empty = str(row2[0]).strip() == "" if row2 else True
            row2_values = [str(v).strip() for v in row2[1:] if str(v).strip()] if len(row2) > 1 else []
            def _is_year_like(s):
                if not s:
                    return False
                if s.isdigit() and 1900 <= int(s) <= 2100:
                    return True
                if '/' in s:
                    parts = s.split('/')
                    return all(p.isdigit() and 1900 <= int(p) <= 2100 for p in parts if p)
                return False
            row2_all_years = (
                len(row2_values) > 0 and
                all(_is_year_like(v) for v in row2_values)
            )
            if row1_col0_empty and row1_has_content and row2_col0_empty and row2_all_years:
                has_metadata = True
                units = raw_rows[1]
                years = raw_rows[2]

    if has_metadata:
        data_rows = raw_rows[3:]
    else:
        data_rows = raw_rows[1:]
        
    _PD_RE = re.compile(r'^\d+\.\d{1,2}$')

    def _fix_dot_decimal(val: str) -> str:
        s = str(val).strip()
        if _PD_RE.match(s):
            return s.replace('.', ',', 1)
        return val

    records = []
    for r in data_rows:
        record = {}
        for idx, h in enumerate(headers):
            val = r[idx] if idx < len(r) else ""
            if idx > 0 and val:
                val = _fix_dot_decimal(val)
            record[h] = val
        records.append(record)
        
    return headers, records, units, years

@app.post("/api/tables/{table_id}/load")
def load_table_csv(table_id: int, db: Session = Depends(get_db), admin: dict = Depends(require_admin)):
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Table not found")
    
    # Hapus row yang sudah ada jika re-load
    db.query(models.TableRow).filter(models.TableRow.table_id == table_id).delete()
    
    try:
        safe_path = get_safe_windows_path(table.csv_path)
        headers, records, units, years = parse_csv_for_db(safe_path)
        
        # Simpan metadata kolom langsung ke database
        table.headers = headers
        table.units = units
        table.years = years
        
        anomaly_count = 0
        for row_idx, record in enumerate(records):
            normalize_record_first_col(record, headers)
            is_anomaly = False
            # Deteksi anomali: hanya tandai jika mengandung "?"
            for key, val in record.items():
                str_val = str(val).strip()
                if "?" in str_val or str_val == "":
                    is_anomaly = True
                    break
            
            if is_anomaly:
                anomaly_count += 1
                
            db_row = models.TableRow(table_id=table.id, data=record, is_anomaly=is_anomaly, sort_order=row_idx)
            db.add(db_row)
        db.commit()
        return {"message": f"Loaded {len(records)} rows successfully. Found {anomaly_count} anomalies."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading CSV: {str(e)}")

# ===== PENCARIAN TABEL =====
@app.get("/api/tables/search")
def search_tables(
    q: str = "",
    year: int = None,
    document_id: int = None,
    bab: int = None,
    limit: int = 50,
    db: Session = Depends(get_db)
):
    """Search tables by keyword or numbering. Supports document_id and bab filters."""
    try:
        query = db.query(
            models.ExtractedTable,
            models.Document.year,
            models.Document.filename
        ).join(
            models.Document,
            models.ExtractedTable.document_id == models.Document.id
        )

        if document_id:
            query = query.filter(models.ExtractedTable.document_id == document_id)

        if bab is not None:
            query = query.filter(
                (models.ExtractedTable.table_name.ilike(f"Tabel {bab}.%")) |
                (models.ExtractedTable.table_name.ilike(f"Tabel_{bab}.%")) |
                (models.ExtractedTable.table_name.ilike(f"{bab}.%")) |
                (models.ExtractedTable.table_name.ilike(f"Tabel {bab}-%")) |
                (models.ExtractedTable.table_name.ilike(f"Tabel_{bab}-%"))
            )

        if q:
            kw = q.strip().lower()
            # Clean common prefixes like "tabel " or "tabel_" if user typed it
            kw_clean = re.sub(r'^(tabel[\s_]*)', '', kw)
            
            # If search term is a number/dot pattern, e.g. "2.1.1" or "2"
            if re.match(r'^[\d.]+$', kw_clean):
                query = query.filter(
                    (models.ExtractedTable.table_name.ilike(f"Tabel {kw_clean}%")) |
                    (models.ExtractedTable.table_name.ilike(f"Tabel_{kw_clean}%")) |
                    (models.ExtractedTable.table_name.ilike(f"%{kw_clean}%"))
                )
            else:
                query = query.filter(
                    models.ExtractedTable.table_name.ilike(f"%{kw}%")
                )

        if year:
            query = query.filter(models.Document.year == year)

        results = query.order_by(models.ExtractedTable.id).limit(limit).all()

        tables_out = []
        for t, doc_year, doc_name in results:
            if not t:
                continue
            table_name_str = t.table_name or ""
            # Extract chapter number safely
            m = re.search(r'Tabel[\s_]*(\d+)', table_name_str, re.IGNORECASE)
            bab_num = int(m.group(1)) if m else None
            tables_out.append({
                "id": t.id,
                "table_name": table_name_str,
                "document_id": t.document_id,
                "document_year": doc_year,
                "document_name": doc_name,
                "bab_num": bab_num
            })

        return {"tables": tables_out, "total": len(tables_out)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database search failed: {str(e)}")


@app.get("/api/tables/{table_id}")
def get_table_info(table_id: int, db: Session = Depends(get_db)):
    """Get table basic info including whether DB data exists."""
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(404, "Table not found")
    doc = db.query(models.Document).filter(models.Document.id == table.document_id).first()
    has_db = db.query(models.TableRow).filter(models.TableRow.table_id == table_id).first() is not None
    return {
        "id": table.id,
        "table_name": table.table_name or "",
        "csv_path": table.csv_path or "",
        "document_id": table.document_id,
        "document_name": doc.filename if doc else "",
        "document_year": doc.year if doc else None,
        "has_db_data": has_db
    }

@app.put("/api/tables/{table_id}/db_rows")
def save_db_rows(table_id: int, payload: dict, db: Session = Depends(get_db), admin: dict = Depends(require_admin)):
    """Batch update database rows from edit mode."""
    rows = payload.get("rows", [])
    for row_data in rows:
        row_id = row_data.get("id")
        data = row_data.get("data", {})
        is_anomaly = row_data.get("is_anomaly", False)
        if row_id:
            row = db.query(models.TableRow).filter(models.TableRow.id == row_id, models.TableRow.table_id == table_id).first()
            if row:
                row.data = data
                row.is_anomaly = is_anomaly
    db.commit()
    return {"message": f"{len(rows)} baris tersimpan"}

# Get table data from DB with CSV header metadata (unit, year) for unified rendering
@app.get("/api/tables/{table_id}/data")
def get_table_data(table_id: int, db: Session = Depends(get_db)):
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    rows = db.query(models.TableRow).filter(models.TableRow.table_id == table_id).order_by(models.TableRow.sort_order.asc(), models.TableRow.id.asc()).all()
    
    headers = get_table_headers(db, table)
    units = table.units if table and table.units else ([""] * len(headers) if headers else [])
    years = table.years if table and table.years else ([""] * len(headers) if headers else [])
            
    return {
        "headers": headers, 
        "units": units,
        "years": years,
        "rows": [{"id": r.id, "data": r.data, "is_anomaly": r.is_anomaly} for r in rows]
    }

@app.put("/api/data/{row_id}")
def update_row_data(row_id: int, payload: dict, db: Session = Depends(get_db), admin: dict = Depends(require_admin)):
    row = db.query(models.TableRow).filter(models.TableRow.id == row_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Row not found")
    row.data = payload.get("data", {})
    db.commit()
    log_activity(db, "edit_row", f"row_id={row_id}", {"table_id": row.table_id})
    return {"message": "Updated successfully"}

@app.put("/api/data/{row_id}/safe")
def mark_row_safe(row_id: int, db: Session = Depends(get_db), admin: dict = Depends(require_admin)):
    row = db.query(models.TableRow).filter(models.TableRow.id == row_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Row not found")
    row.is_anomaly = False
    db.commit()
    return {"message": "Row marked as safe"}

@app.put("/api/tables/{table_id}/safe-all")
def mark_all_rows_safe(table_id: int, db: Session = Depends(get_db), admin: dict = Depends(require_admin)):
    db.query(models.TableRow).filter(models.TableRow.table_id == table_id).update({"is_anomaly": False})
    db.commit()
    log_activity(db, "safe_anomaly", f"table_id={table_id}")
    return {"message": "All rows marked as safe"}

@app.delete("/api/data/{row_id}")
def delete_row(row_id: int, db: Session = Depends(get_db), admin: dict = Depends(require_admin)):
    row = db.query(models.TableRow).filter(models.TableRow.id == row_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Row not found")
    table_id = row.table_id
    db.delete(row)
    db.commit()
    log_activity(db, "delete_row", f"row_id={row_id}", {"table_id": table_id})
    return {"message": "Deleted successfully"}

def get_clean_chapter_name(level1: str) -> str:
    CHAPTER_NAMES = {
        "1": "Geografi dan Iklim",
        "2": "Pemerintahan",
        "3": "Penduduk dan Ketenagakerjaan",
        "4": "Sosial dan Kesejahteraan Rakyat",
        "5": "Pertanian, Kehutanan, dan Perikanan",
        "6": "Industri, Pertambangan, Energi, dan Air",
        "7": "Pariwisata",
        "8": "Transportasi dan Komunikasi",
        "9": "Koperasi dan Usaha Mikro Kecil Menengah (UMKM)",
        "10": "Pengeluaran dan Konsumsi Penduduk",
        "11": "Perdagangan",
        "12": "Pendapatan Regional",
        "13": "Perbandingan Regional / Antar Wilayah"
    }
    return CHAPTER_NAMES.get(str(level1).strip(), f"Bab {level1}")

def get_clean_table_name(table_name: str) -> str:
    if not table_name:
        return ""
    # 1. Hapus awalan nomor tabel (contoh: "Tabel 4.4.2 - " atau "Tabel 4.4.2 ")
    name = re.sub(r'^(?:Tabel[\s_]*\d+(?:\.\d+)*\s*(?:-\s*|:\s*|)\s*)', '', table_name, flags=re.IGNORECASE)
    # 2. Hapus .csv di akhir jika ada
    name = re.sub(r'\.csv$', '', name, flags=re.IGNORECASE)
    # 3. Hapus referensi halaman seperti (Hal 46), (Hal 47, 48), (Halaman 12), dll.
    name = re.sub(r'\s*\((?:Hal|Halaman|hlm)[\s\d,\-–—\.\?]+\)', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s*\(\s*\d+[\s,\d\-–—\.]*\)\s*$', '', name)
    # 4. Hapus 'Tahun 2022', 'Pada Tahun 2021-2022', 'Year 2025' atau sisa 'Tahun' di ujung akhir
    year_token = r'(?:19|20)\d{2}[*\d]?'
    year_conn = r'(?:\s*(?:[-–—/]|dan|and|sd|s/d|to|,)\s*' + year_token + r')*'
    name = re.sub(r'[,.\s]+(?:(?:pada|di)\s+)?(?:tahun|years?)\s*(?:' + year_token + year_conn + r'.*)?$', '', name, flags=re.IGNORECASE)
    # 5. Hapus tahun langsung jika tanpa kata 'tahun', misal ', 2022' atau ' 2021-2025'
    name = re.sub(r'[,.\s]+' + year_token + year_conn + r'.*$', '', name, flags=re.IGNORECASE)
    # 6. Hapus sisa kata 'Tahun' / 'Year' jika masih ada di ujung akhir
    name = re.sub(r'[,.\s]+(?:(?:pada|di)\s+)?(?:tahun|years?)\s*$', '', name, flags=re.IGNORECASE)
    return re.sub(r'[,.\-\s–—]+$', '', name).strip()

@app.get("/api/tables/{table_id}/excel")
@app.get("/api/tables/{table_id}/export_excel")
def download_table_excel(table_id: int, db: Session = Depends(get_db)):
    """Unduh data 1 tabel dalam format Microsoft Excel (.xlsx) asli ber-styling rapi, auto column width, dan cell border."""
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Tabel tidak ditemukan")
    
    headers = get_table_headers(db, table)
    if not headers:
        raise HTTPException(status_code=404, detail="Tabel tidak memiliki data kolom")
    
    units = list(table.units) if table.units else [""] * len(headers)
    years = list(table.years) if table.years else [""] * len(headers)
    while len(units) < len(headers):
        units.append("")
    while len(years) < len(headers):
        years.append("")

    all_rows = db.query(models.TableRow).filter(models.TableRow.table_id == table_id).order_by(models.TableRow.sort_order.asc(), models.TableRow.id.asc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = re.sub(r'[\\/?*\[\]:]', ' ', (table.table_name or "Tabel"))[:31]
    ws.views.sheetView[0].showGridLines = True

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # 1. Header Baris 1: Nama Kolom
    ws.append(headers)
    ws.row_dimensions[1].height = 26
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1E40AF')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # 2. Header Baris 2: Satuan (jika ada)
    has_units = any(str(u).strip() for u in units)
    if has_units:
        ws.append(units)
        u_row = ws.max_row
        ws.row_dimensions[u_row].height = 20
        for cell in ws[u_row]:
            cell.font = Font(name='Segoe UI', size=9, italic=True, color='475569')
            cell.fill = PatternFill('solid', fgColor='F1F5F9')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # 3. Baris Data
    for r in all_rows:
        row_vals = []
        for h in headers:
            val_str = str(r.data.get(h, "")).strip() if isinstance(r.data, dict) else ""
            clean_num = val_str.replace('.', '').replace(',', '.')
            try:
                val_num = float(clean_num)
                row_vals.append(val_num)
            except ValueError:
                row_vals.append(val_str)
        ws.append(row_vals)
        r_idx = ws.max_row
        ws.row_dimensions[r_idx].height = 20
        for col_idx, cell in enumerate(ws[r_idx], 1):
            cell.font = Font(name='Segoe UI', size=10, color='1E293B')
            cell.border = thin_border
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0.00' if isinstance(cell.value, float) and not cell.value.is_integer() else '#,##0'
            else:
                cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'center', vertical='center')

    # 4. Auto Column Width (Lebar Otomatis berdasarkan panjang teks + padding)
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                s = str(cell.value)
                w = sum(2 if ord(ch) > 127 else 1 for ch in s)
                if w > max_len:
                    max_len = w
        ws.column_dimensions[col_letter].width = max(min(max_len + 4, 60), 14)

    ws.freeze_panes = 'A3' if has_units else 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    clean_title = re.sub(r'[\\/:*?"<>|]', '_', (table.table_name or f"tabel_{table_id}")).strip()
    clean_title = clean_title.encode("ascii", "ignore").decode("ascii").strip() or f"tabel_{table_id}"
    filename = f"{clean_title}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.get("/api/tables/{table_id}/csv")
def download_table_csv(table_id: int, db: Session = Depends(get_db)):
    """Unduh data 1 tabel dalam format CSV dengan UTF-8 BOM agar terbaca langsung terpisah kolom di Excel."""
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Tabel tidak ditemukan")
    
    headers = get_table_headers(db, table)
    if not headers:
        raise HTTPException(status_code=404, detail="Tabel tidak memiliki data kolom")
    
    units = table.units or [""] * len(headers)
    years = table.years or [""] * len(headers)
    all_rows = db.query(models.TableRow).filter(models.TableRow.table_id == table_id).order_by(models.TableRow.sort_order.asc(), models.TableRow.id.asc()).all()
    
    buf = io.StringIO()
    # Tulis UTF-8 BOM dan deklarasi 'sep=,' agar Excel mengenali delimiter koma secara otomatis di regional setting manapun
    buf.write('\ufeffsep=,\r\n')
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerow(units)
    writer.writerow(years)
    for r in all_rows:
        writer.writerow([r.data.get(h, "") for h in headers])
    buf.seek(0)
    
    filename = re.sub(r'[\\/:*?"<>|]', '_', (table.table_name or f"tabel_{table_id}")) + ".csv"
    filename = filename.encode("ascii", "ignore").decode("ascii").strip()
    if not filename:
        filename = f"tabel_{table_id}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@app.get("/api/tables/{table_id}/csv_preview")
def preview_table_csv(table_id: int, db: Session = Depends(get_db)):
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Tabel tidak ditemukan")
    
    try:
        orig_headers = get_table_headers(db, table)
        if not orig_headers:
            return {"headers": [], "units": [], "years": [], "rows": []}
        
        units = list(table.units) if table.units else [""] * len(orig_headers)
        years = list(table.years) if table.years else [""] * len(orig_headers)
        
        # Normalize units "Persen" -> "%"
        units = ["%" if str(u).lower() in ["persen", "persentase", "percent"] else u for u in units]
        
        # Bersihkan nama kolom bilingual
        headers = [clean_bilingual_header(h) for h in orig_headers]
        
        all_rows = db.query(models.TableRow).filter(models.TableRow.table_id == table_id).order_by(models.TableRow.sort_order.asc(), models.TableRow.id.asc()).all()
        data_rows = [[r.data.get(h, "") for h in orig_headers] for r in all_rows]
        
        return {
            "table_id": table.id,
            "table_name": table.table_name,
            "headers": headers,
            "orig_headers": orig_headers,
            "units": units, 
            "years": years, 
            "rows": data_rows,
            "row_ids": [r.id for r in all_rows],
            "is_anomalies": [bool(r.is_anomaly) for r in all_rows]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==========================================================
# ADMIN DATABASE ENDPOINTS
# ==========================================================

@app.get("/api/admin/tables")
def admin_get_tables(db: Session = Depends(get_db), admin: dict = Depends(require_admin)):
    # Ambil semua tabel beserta info tahun & dokumennya
    tables = db.query(
        models.ExtractedTable.id, 
        models.ExtractedTable.table_name, 
        models.ExtractedTable.csv_path,
        models.Document.year,
        models.Document.filename
    ).join(models.Document, models.ExtractedTable.document_id == models.Document.id).all()
    
    result = []
    for t in tables:
        # Hitung jumlah baris data yang SUDAH di-load ke database (TableRow)
        row_count = db.query(models.TableRow).filter(models.TableRow.table_id == t.id).count()
        table_name_str = t.table_name or ""
        m = re.search(r'Tabel[\s_]*(\d+)', table_name_str, re.IGNORECASE)
        result.append({
            "id": t.id,
            "table_name": table_name_str,
            "csv_path": t.csv_path,
            "year": t.year,
            "db_rows": row_count,
            "document_name": t.filename,
            "bab_num": int(m.group(1)) if m else None,
            "has_db_data": row_count > 0
        })
    
    return result

@app.post("/api/admin/clear-loaded-data")
def clear_loaded_data(db: Session = Depends(get_db), admin: dict = Depends(require_admin)):
    """Clear all TableRow data from the database, but keep ExtractedTable and Document records."""
    try:
        db.query(models.TableRow).delete()
        db.commit()
        return {"message": "All loaded table rows have been cleared successfully."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to clear data: {str(e)}")

@app.post("/api/documents/{doc_id}/load-all")
def load_all_document_tables(doc_id: int, db: Session = Depends(get_db), admin: dict = Depends(require_admin)):
    tables = db.query(models.ExtractedTable).filter(models.ExtractedTable.document_id == doc_id).all()
    loaded_count = 0
    errors = 0
    for t in tables:
        try:
            db.query(models.TableRow).filter(models.TableRow.table_id == t.id).delete()
            safe_path = get_safe_windows_path(t.csv_path)
            headers, records, units, years = parse_csv_for_db(safe_path)
            t.headers = headers
            t.units = units
            t.years = years
            for row_idx, record in enumerate(records):
                is_anomaly = False
                for key, val in record.items():
                    str_val = str(val).strip()
                    if "?" in str_val:
                        is_anomaly = True
                        break
                db_row = models.TableRow(table_id=t.id, data=record, is_anomaly=is_anomaly, sort_order=row_idx)
                db.add(db_row)
            loaded_count += 1
        except Exception:
            errors += 1
    db.commit()
    log_activity(db, "reload_all", f"doc_id={doc_id}", {"loaded": loaded_count, "errors": errors})
    return {"message": f"Berhasil me-load {loaded_count} tabel ke database. Gagal: {errors} tabel."}

@app.post("/api/documents/{doc_id}/bab/{bab_num}/load-all")
def load_all_chapter_tables(doc_id: int, bab_num: int, db: Session = Depends(get_db), admin: dict = Depends(require_admin)):
    tables = db.query(models.ExtractedTable).filter(models.ExtractedTable.document_id == doc_id).all()
    loaded_count = 0
    errors = 0

    for t in tables:
        match = re.search(r'Tabel[\s_]*(\d+)', t.table_name, re.IGNORECASE)
        if match and int(match.group(1)) == bab_num:
            try:
                db.query(models.TableRow).filter(models.TableRow.table_id == t.id).delete()
                safe_path = get_safe_windows_path(t.csv_path)
                headers, records, units, years = parse_csv_for_db(safe_path)
                t.headers = headers
                t.units = units
                t.years = years
                for row_idx, record in enumerate(records):
                    is_anomaly = False
                    for key, val in record.items():
                        str_val = str(val).strip()
                        if "?" in str_val:
                            is_anomaly = True
                            break
                    db_row = models.TableRow(table_id=t.id, data=record, is_anomaly=is_anomaly, sort_order=row_idx)
                    db.add(db_row)
                loaded_count += 1
            except Exception:
                errors += 1
    db.commit()
    log_activity(db, "reload_chapter", f"bab {bab_num}", {"doc_id": doc_id, "loaded": loaded_count, "errors": errors})
    return {"message": f"Berhasil me-load {loaded_count} tabel Bab {bab_num} ke database. Gagal: {errors} tabel."}

# ===== FIX TRUNCATED TABLE NAMES =====
@app.post("/api/admin/fix-table-names")
def fix_truncated_table_names(db: Session = Depends(get_db), admin: dict = Depends(require_admin)):
    """
    Memperbaiki nama tabel yang terpotong:
    1. Cari metadata.json di seluruh folder ekstraksi untuk ambil judul LENGKAP
    2. Fallback: pakai nama file CSV jika metadata tidak ditemukan
    """
    import os
    import glob
    import unicodedata
    
    extract_root = os.path.join(os.path.expanduser("~"), "BPS_Data", "hasil_ekstraksi_web")
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    outputs_dir = os.path.join(project_root, "outputs")
    
    full_title_map = {}
    
    meta_files = glob.glob(os.path.join(extract_root, "**", "metadata.json"), recursive=True)
    if os.path.exists(outputs_dir):
        outputs_meta = glob.glob(os.path.join(outputs_dir, "**", "metadata.json"), recursive=True)
        meta_files.extend(outputs_meta)
    
    for meta_path in meta_files:
        try:
            with open(meta_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            tm = data.get("title_mapping", {})
            for csv_fn, full_title in tm.items():
                csv_fn_norm = csv_fn.replace(" __SLASH__ ", "/").replace("__SLASH__", "/")
                full_title_norm = full_title.replace(" __SLASH__ ", "/").replace("__SLASH__", "/")
                if csv_fn_norm not in full_title_map or len(full_title_norm) > len(full_title_map[csv_fn_norm]):
                    full_title_map[csv_fn_norm] = full_title_norm
        except Exception:
            continue
    
    tables = db.query(models.ExtractedTable).all()
    fixed = []
    skipped = []
    for t in tables:
        old_name = t.table_name or ""
        csv_path = t.csv_path or ""
        csv_file = os.path.basename(csv_path)
        
        if not csv_file:
            skipped.append({"id": t.id, "reason": "no_csv_path"})
            continue
        
        csv_file_norm = csv_file.replace(" __SLASH__ ", "/").replace("__SLASH__", "/")
        csv_noext = csv_file_norm.rsplit(".", 1)[0].strip() if csv_file_norm.lower().endswith(".csv") else csv_file_norm
        full_title = full_title_map.get(csv_file_norm) or full_title_map.get(csv_noext)
        
        if full_title:
            new_name = full_title.replace(" __SLASH__ ", "/").replace("__SLASH__", "/").strip()
        else:
            csv_name = csv_file.rsplit(".", 1)[0] if csv_file.lower().endswith(".csv") else csv_file
            csv_name_clean = csv_name.replace(" __SLASH__ ", "/").replace("__SLASH__", "/")
            if len(csv_name_clean) <= len(old_name) or csv_name_clean == old_name:
                skipped.append({"id": t.id, "reason": "no_longer_name"})
                continue
            new_name = csv_name_clean
        
        if unicodedata.normalize('NFKC', new_name) == unicodedata.normalize('NFKC', old_name):
            skipped.append({"id": t.id, "reason": "same_name"})
            continue
        
        t.table_name = new_name
        fixed.append({"id": t.id, "old": old_name, "new": new_name, "source": "metadata" if full_title else "csv_filename"})
    
    db.commit()
    return {
        "message": f"{len(fixed)} nama tabel diperbaiki, {len(skipped)} dilewati",
        "fixed": len(fixed),
        "skipped": len(skipped),
        "details": fixed[:50]
    }

# ===== TABLE NEIGHBORS =====
@app.get("/api/tables/{table_id}/neighbors")
def get_table_neighbors(table_id: int, db: Session = Depends(get_db)):
    """Get next and previous table IDs within the same document."""
    table = db.query(models.ExtractedTable).filter(models.ExtractedTable.id == table_id).first()
    if not table:
        raise HTTPException(404, "Table not found")
    
    siblings = db.query(models.ExtractedTable).filter(
        models.ExtractedTable.document_id == table.document_id
    ).all()
    
    def natural_sort_key(t):
        name = t.table_name or ""
        match = re.search(r'(\d+(?:\.\d+)+)', name)
        if match:
            try:
                parts = [int(p) for p in match.group(1).split('.')]
                return (0, parts, name.lower())
            except ValueError:
                pass
        return (1, [], name.lower())
        
    siblings.sort(key=natural_sort_key)
    
    prev_id = None
    next_id = None
    curr_idx = -1
    for i, s in enumerate(siblings):
        if s.id == table_id:
            curr_idx = i
            if i > 0:
                prev_id = siblings[i - 1].id
            if i < len(siblings) - 1:
                next_id = siblings[i + 1].id
            break
    
    return {
        "prev_id": prev_id,
        "next_id": next_id,
        "prev_name": siblings[curr_idx - 1].table_name if curr_idx > 0 else None,
        "next_name": siblings[curr_idx + 1].table_name if curr_idx < len(siblings) - 1 else None,
        "current_index": curr_idx,
        "total_in_doc": len(siblings)
    }

# =====================================================================
# CATCH-ALL ROUTE — SPA fallback untuk URL yang tidak dikenali
# =====================================================================
@app.get("/{path:path}")
def catch_all(request: Request, path: str):
    if path.startswith("api/"):
        raise HTTPException(status_code=404, detail="Endpoint tidak ditemukan.")
    return templates.TemplateResponse(request=request, name="404.html", status_code=404)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
